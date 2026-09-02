import json, os
from openpyxl import load_workbook

sales_path='/Users/iant1359/Downloads/So_chi_tiet_ban_hang 01.01.2023 đến 09.07.26.xlsx'
inv_path='/Users/iant1359/Downloads/Tong_hop_ton_kho 09.07.xlsx'
out='/tmp/amis-sheet/clean_data.json'

def iso(v):
    return v.strftime('%Y-%m-%d') if hasattr(v,'strftime') else v

sw=load_workbook(sales_path,read_only=True,data_only=True).active
headers=[str(x) if x is not None else '' for x in next(sw.iter_rows(min_row=4,max_row=4,values_only=True))]
sales=[]; sales_removed=0
for r in sw.iter_rows(min_row=5,values_only=True):
    row=[iso(x) for x in r[:22]]
    note=str(r[12]).strip() if r[12] is not None else ''
    debit=str(r[18]).strip() if r[18] is not None else ''
    credit=str(r[19]).strip() if r[19] is not None else ''
    is_rev=debit in {'5111','5112','5113'} or credit in {'5111','5112','5113'}
    if is_rev and note.lower() != 'không phải revenue': sales.append(row)
    else: sales_removed += 1

iw=load_workbook(inv_path,read_only=True,data_only=True).active
inv_headers=['Tên kho','Mã kho','Mã hàng','Tên hàng','ĐVT','Cuối kỳ - Số lượng','Cuối kỳ - Giá trị','Nhóm VTHH','KT note']
excluded={'Kho Bình Phú Hàng Lỗi (Kho ảo)','Kho Bình Phú (Kho lỗi)','Kho Chị Kathy','Kho chưa xuất hóa đơn'}
inv=[]; inv_removed=0
for r in iw.iter_rows(min_row=6,values_only=True):
    wh=str(r[0]).strip() if r[0] is not None else ''
    note=str(r[14]).strip() if r[14] is not None else ''
    if wh in excluded or note == 'Loại khỏi tồn kho': inv_removed += 1; continue
    inv.append([r[0],r[1],r[2],r[3],r[4],r[11],r[12],r[13],r[14]])

summary={
 'sales_total_rows': sw.max_row-4, 'sales_kept_rows':len(sales), 'sales_removed_rows':sales_removed,
 'sales_revenue': sum((x[17] or 0) for x in sales if isinstance(x[17],(int,float))),
 'inventory_total_rows': iw.max_row-5, 'inventory_kept_rows':len(inv), 'inventory_removed_rows':inv_removed,
 'inventory_qty': sum((x[5] or 0) for x in inv if isinstance(x[5],(int,float))),
 'inventory_value': sum((x[6] or 0) for x in inv if isinstance(x[6],(int,float))),
 'excluded_warehouses': sorted(excluded),
}
with open(out,'w',encoding='utf-8') as f: json.dump({'sales_headers':headers[:22],'sales':sales,'inventory_headers':inv_headers,'inventory':inv,'summary':summary},f,ensure_ascii=False)
print(out)
