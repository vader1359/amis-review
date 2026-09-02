import json
from collections import defaultdict
from openpyxl import load_workbook
IN='input'; OUT='/tmp/amis-sheet/psi_data.json'
def cv(x): return x.strftime('%Y-%m-%d') if hasattr(x,'strftime') else x
def norm(x): return str(x).strip().upper() if x not in (None,'') else ''
pw=load_workbook(IN+'/CRM_Product_10.07.2026_08.32.37_482.xlsx',read_only=True,data_only=True).active
products=[]; pmap={}
for r in pw.iter_rows(min_row=2,values_only=True):
    if not r[0]: continue
    z=[r[0],r[2],r[6],r[7],r[21] if len(r)>21 else None,r[22] if len(r)>22 else None,r[13]]; products.append(z); pmap[str(r[0]).strip()]=z
known_aliases={'USMUS-11219.3','USMUS-14294.3','USMUS-14298.3','USMUS-14300.3','USMUS-14296.3','USMUS-11207.3','USMUS-16853.3','USMUS-11211.3','USMUS-16845.3'}
sw=load_workbook(IN+'/So_chi_tiet_ban_hang 01.01.2023 đến 09.07.26.xlsx',read_only=True,data_only=True).active; rh=list(next(sw.iter_rows(min_row=4,max_row=4,values_only=True)))[:22]; revenue=[]; mis=[]
for r in sw.iter_rows(min_row=5,values_only=True):
    if (norm(r[18]) in {'5111','5112','5113'} or norm(r[19]) in {'5111','5112','5113'}) and norm(r[12])!='KHÔNG PHẢI REVENUE':
        revenue.append([cv(x) for x in r[:22]]); code=str(r[10]).strip() if r[10] else ''
        if code and code not in pmap and code not in known_aliases: mis.append([code,r[11],'Revenue'])
iw=load_workbook(IN+'/Tong_hop_ton_kho 09.07.xlsx',read_only=True,data_only=True).active; inv=[]; excluded={'KHO BÌNH PHÚ HÀNG LỖI (KHO ẢO)','KHO BÌNH PHÚ (KHO LỖI)','KHO CHỊ KATHY','KHO CHƯA XUẤT HÓA ĐƠN'}
for r in iw.iter_rows(min_row=6,values_only=True):
    if norm(r[0]) in excluded or norm(r[14])=='LOẠI KHỎI TỒN KHO': continue
    inv.append([r[0],r[1],r[2],r[3],r[4],r[11],r[12],r[13],r[14]]); code=str(r[2]).strip() if r[2] else ''
    if code and code not in pmap and code not in known_aliases: mis.append([code,r[3],'Inventory'])
prew=load_workbook(IN+'/Pre order feedback.xlsx',read_only=True,data_only=True).active; preh=list(next(prew.iter_rows(min_row=1,max_row=1,values_only=True))); pre=[]
for r in prew.iter_rows(min_row=2,values_only=True):
    pre.append([cv(x) for x in r[:10]]); code=str(r[1]).strip() if r[1] else ''
    if code and code not in pmap: mis.append([code,r[2],'Pre-order'])
cw=load_workbook(IN+'/CRM_Saleorder_06.07.2026_09.59.50_041.xlsx',read_only=True,data_only=True)['Bảng hàng hóa']; ch=list(next(cw.iter_rows(min_row=1,max_row=1,values_only=True)))[:15]; crm=[[cv(x) for x in r[:15]] for r in cw.iter_rows(min_row=2,values_only=True)]
tw=load_workbook(IN+'/Target.xlsx',read_only=True,data_only=True).active; target=[[cv(x) for x in r] for r in tw.iter_rows(min_row=2,values_only=True)]
lw=load_workbook(IN+'/0_LOADING LIST DETAIL_New.xlsx',read_only=True,data_only=True)['LDL']; ph=['Mã MISA','Tên hàng hóa','Brand','Category','Sub-category','Số lượng','Giá trị nhập kho','Tình trạng','PRE-ORDER/STOCK','Ngày nhập kho']; purchase=[]
aliases={'USMUS10200':'USMUS10200-CB','USMUS10201':'USMUS10201-CB'}
for r in lw.iter_rows(min_row=5,values_only=True):
    if not any(x not in (None,'',0) for x in r[:79]): continue
    code=str(r[18]).strip() if r[18] not in (None,'') else ''
    name=str(r[19]).replace('\n',' / ') if r[19] else ''
    if not code and 'HILL HOUSE 1' in name.upper() and 'OIL GREEN' in name.upper(): code='CHRCA00016'
    if not code and 'PLATTER 210' in name.upper(): code='LWLFL00029'
    purchase.append([code,name,r[6],r[14],r[15],r[20],r[77],r[1],r[7],cv(r[54])])
    if code and code not in pmap and code not in aliases: mis.append([code,name,'Purchase/PO'])
    if not code and name: mis.append(['',name,'Purchase/PO - thiếu mã MISA'])
seen=set(); mm=[]
for x in mis:
    k=tuple(x)
    if k not in seen: seen.add(k); mm.append(x)
agg=defaultdict(lambda:[0,0,0,0,0])
for r in revenue:
    k=str(r[10]).strip() if r[10] else ''; agg[k][0]+=r[15] or 0 if isinstance(r[15],(int,float)) else 0; agg[k][1]+=r[17] or 0 if isinstance(r[17],(int,float)) else 0
for r in inv:
    k=str(r[2]).strip() if r[2] else ''; agg[k][2]+=r[5] or 0 if isinstance(r[5],(int,float)) else 0; agg[k][3]+=r[6] or 0 if isinstance(r[6],(int,float)) else 0
for r in pre:
    k=str(r[0]).strip() if r[0] else ''; agg[k][4]+=r[5] or 0 if isinstance(r[5],(int,float)) else 0
psi_rows=[]
for p in products:
    a=agg[p[0]]
    if any(a): psi_rows.append([p[0],p[1],p[2],p[3],p[4],p[5],a[0],a[1],a[2],a[3],a[4]])
summary=[['Metric','Value'],['Products',len(products)],['Revenue lines kept',len(revenue)],['Inventory lines kept',len(inv)],['Pre-order lines',len(pre)],['CRM order-item lines',len(crm)],['Purchase/PO lines',len(purchase)],['Unique mismatch rows',len(mm)],['Purchase source','input/0_LOADING LIST DETAIL_New.xlsx / LDL']]
def read_simple(path,sheet,min_row=1,max_cols=25):
    ws=load_workbook(path,read_only=True,data_only=True)[sheet]; it=ws.iter_rows(min_row=min_row,values_only=True); h=list(next(it))[:max_cols]; rr=[[cv(x) for x in r[:max_cols]] for r in it]; return h,rr
crm_h,crm_rows=read_simple(IN+'/CRM_Sale.xlsx','Danh sách',1,25)
ord_h,ord_rows=read_simple(IN+'/CRM_Saleorder_06.07.2026_09.59.50_041.xlsx','Danh sách',1,25)
old_h,old_rows=read_simple('test/PSI_Data 02.07.xlsx','Revenue cũ',6,15)
brand_rows=sorted({(str(p[2] or ''),str(p[6] or '')) for p in products if p[2] or p[6]})
cat_rows=sorted({(str(p[4] or ''),str(p[5] or '')) for p in products if p[4] or p[5]})
with open(OUT,'w',encoding='utf-8') as f: json.dump({'summary':summary,'brand_headers':['Brand','Mã hãng'],'brand':brand_rows,'category_headers':['Category','Sub Category'],'category':cat_rows,'psi_headers':['Mã hàng','Tên hàng','Brand','Loại hàng','Category','Sub Category','SL bán','Doanh số','SL tồn cuối kỳ','GT tồn cuối kỳ','SL pre-order'],'psi_rows':psi_rows,'products':products,'revenue_headers':rh,'revenue':revenue,'inventory_headers':['Tên kho','Mã kho','Mã hàng','Tên hàng','ĐVT','Cuối kỳ - SL','Cuối kỳ - Giá trị','Nhóm VTHH','KT note'],'inventory':inv,'pre_headers':preh,'pre':pre,'purchase_headers':ph,'purchase':purchase,'crm_headers':crm_h,'crm_data':crm_rows,'order_headers':ord_h,'order_data':ord_rows,'old_revenue_headers':old_h,'old_revenue':old_rows,'target_headers':['Brand','Brand Code','Main Showroom','Target 2026'],'target':target,'mismatch_headers':['Mã hàng','Tên hàng','Nguồn'],'mismatch':mm},f,ensure_ascii=False,default=str)
print(json.dumps({'products':len(products),'revenue':len(revenue),'inventory':len(inv),'pre':len(pre),'crm':len(crm),'purchase':len(purchase),'mismatch':len(mm)}))
