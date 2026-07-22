from __future__ import annotations

import io
from collections import defaultdict
from dataclasses import dataclass
from typing import Final

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill


ALIAS: Final = {
    "USMUS-11219.3", "USMUS-14294.3", "USMUS-14298.3", "USMUS-14300.3",
    "USMUS-14296.3", "USMUS-11207.3", "USMUS-16853.3", "USMUS-11211.3",
    "USMUS-16845.3",
}
FileSet = dict[str, bytes]
Cell = str | int | float | None
Summary = dict[str, int | float]


@dataclass(frozen=True, slots=True)
class PsiBuildResult:
    summary: Summary
    issues: list[list[Cell]]
    gaps: list[list[Cell]]
    xlsx: bytes


def norm(value: Cell) -> str:
    return str(value).strip().upper() if value not in (None, "") else ""


def classify(name: str) -> str | None:
    lowered = name.lower()
    for key, words in {
        "product": ["product"],
        "purchase": ["loading", "purchase", "po"],
        "revenue": ["so_chi_tiet", "revenue"],
        "inventory": ["tong_hop_ton_kho", "inventory"],
        "preorder": ["pre order", "pre-order", "preorder"],
        "crm": ["crm_sale", "crm"],
        "target": ["target"],
    }.items():
        if any(word in lowered for word in words):
            return key
    return None


def build(files: FileSet) -> PsiBuildResult:
    by: dict[str, bytes] = {}
    source_names: dict[str, str] = {}
    gaps: list[list[Cell]] = []
    for name, payload in files.items():
        source_class = classify(name)
        if source_class is None:
            continue
        if source_class not in by or (source_class == "purchase" and "loading" in name.lower()):
            by[source_class] = payload
            source_names[source_class] = name
    required: Final = {
        "product": ["Mã hàng hóa", "Nguồn gốc", "Category", "Sub Category"],
        "purchase": ["SỐ PO", "NGÀY PO", "SỐ DH", "MÃ MISA", "SL", "NGÀY NHẬP KHO", "F.O.C"],
        "revenue": ["Mã hàng", "Doanh số bán", "TK Nợ", "TK Có"],
        "inventory": ["Tên kho", "Mã kho", "Mã hàng", "Cuối kỳ"],
        "preorder": ["PRODUCT ID", "QUANTITY SOLD", "ĐH"],
        "crm": ["PRODUCT ID", "SALE ORDER"],
        "target": ["TARGET 2026"],
    }
    product: dict[str, tuple[Cell, ...]] = {}
    issues: list[list[Cell]] = []
    if "product" in by:
        worksheet = load_workbook(io.BytesIO(by["product"]), read_only=True, data_only=True).active
        header = [norm(value) for value in next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True))]
        for column in required["product"]:
            if not any(norm(value) == norm(column) or norm(column) in norm(value) for value in header):
                gaps.append(["Product", column, "Không tìm thấy cột bắt buộc"])
        for row in worksheet.iter_rows(min_row=2, values_only=True):
            if row[0]:
                product[str(row[0]).strip()] = row
    revenue: list[tuple[Cell, ...]] = []
    if "revenue" in by:
        worksheet = load_workbook(io.BytesIO(by["revenue"]), read_only=True, data_only=True).active
        header = [norm(value) for value in next(worksheet.iter_rows(min_row=4, max_row=4, values_only=True))]
        for column in required["revenue"]:
            if not any(norm(column) in value for value in header):
                gaps.append(["Revenue", column, "Không tìm thấy cột bắt buộc"])
        for row_number, row in enumerate(worksheet.iter_rows(min_row=5, values_only=True), start=5):
            if len(row) > 19 and (norm(row[18]) in {"5111", "5112", "5113"} or norm(row[19]) in {"5111", "5112", "5113"}) and norm(row[12]) != "KHÔNG PHẢI REVENUE":
                revenue.append(row)
                code = str(row[10]).strip() if row[10] else ""
                if code and "product" in by and code not in product and code not in ALIAS:
                    issues.append(["Revenue", source_names["revenue"], worksheet.title, row_number, code, row[11] or "", "Missing Product mapping"])
    inventory: list[tuple[Cell, ...]] = []
    quantity = value = 0
    if "inventory" in by:
        worksheet = load_workbook(io.BytesIO(by["inventory"]), read_only=True, data_only=True).active
        excluded = {"KHO BÌNH PHÚ HÀNG LỖI (KHO ẢO)", "KHO BÌNH PHÚ (KHO LỖI)", "KHO CHỊ KATHY", "KHO CHƯA XUẤT HÓA ĐƠN"}
        for row in worksheet.iter_rows(min_row=6, values_only=True):
            if norm(row[0]) in excluded or norm(row[14]) == "LOẠI KHỎI TỒN KHO":
                continue
            inventory.append(row)
            quantity += row[11] if isinstance(row[11], (int, float)) else 0
            value += row[12] if isinstance(row[12], (int, float)) else 0
    purchase: list[tuple[Cell, ...]] = []
    purchase_excluded: list[list[Cell]] = []
    foc = excluded_po = 0
    if "purchase" in by:
        worksheet = load_workbook(io.BytesIO(by["purchase"]), read_only=True, data_only=True)["LDL"]
        header = [norm(value) for value in next(worksheet.iter_rows(min_row=4, max_row=4, values_only=True))]
        for column in required["purchase"]:
            if not any(norm(column) in value for value in header):
                gaps.append(["Purchase/PO", column, "Không tìm thấy cột bắt buộc"])
        for row_number, row in enumerate(worksheet.iter_rows(min_row=5, values_only=True), start=5):
            if not any(value not in (None, "", 0) for value in row[:79]):
                continue
            kind = norm(row[7]); flag = norm(row[26]); code = str(row[18]).strip() if row[18] else ""; name = str(row[19]).replace("\n", " / ") if row[19] else ""
            if flag in {"F.O.C", "KHÔNG KHAI, PHÂN BỔ VÀO MÃ KHÁC"} or kind in {"CLAIM", "CAMPAIGN", "MARKETING F.O.C", "MARKETING MATERIAL", "SHOWROOM", "FDS", "TEAM DỰ ÁN"}:
                excluded_po += 1; foc += flag == "F.O.C"; purchase_excluded.append([code, name, "F.O.C" if flag else kind, row[1], row[7], row[26]]); continue
            purchase.append(row)
            if not code and name:
                issues.append(["Purchase/PO", source_names["purchase"], worksheet.title, row_number, "", name, "Missing MISA code"])
            elif code and "product" in by and code not in product and code not in {"USMUS10200", "USMUS10201"}:
                issues.append(["Purchase/PO", source_names["purchase"], worksheet.title, row_number, code, name, "Missing Product mapping"])
    workbook = Workbook(); summary_sheet = workbook.active; summary_sheet.title = "PSI Summary"; summary_sheet.append(["Metric", "Value"])
    rows = [("Product master rows", len(product)), ("Revenue lines kept", len(revenue)), ("Inventory rows kept", len(inventory)), ("Inventory quantity", quantity), ("Inventory value", value), ("Purchase rows used", len(purchase)), ("PO rows excluded", excluded_po), ("FOC rows excluded", foc), ("Mismatch issues", len(issues))]
    for row in rows: summary_sheet.append(row)
    mismatch = workbook.create_sheet("Mismatch"); mismatch.append(["Source", "File", "Sheet", "Row", "Code", "Description", "Issue"]); [mismatch.append(row) for row in issues]
    data_gaps = workbook.create_sheet("Data gaps"); data_gaps.append(["Source", "Column / information", "Action required"]); [data_gaps.append(row) for row in gaps]
    excluded_sheet = workbook.create_sheet("PO excluded"); excluded_sheet.append(["Mã MISA", "Tên hàng", "Lý do loại", "Tình trạng", "Phân loại", "F.O.C"]); [excluded_sheet.append(row) for row in purchase_excluded]
    product_sheet = workbook.create_sheet("PSI by Product"); product_sheet.append(["Mã hàng", "Tên hàng", "Brand", "Category", "Sub Category", "SL bán", "Doanh số", "SL tồn", "GT tồn"])
    aggregate: defaultdict[str, list[int | float]] = defaultdict(lambda: [0, 0, 0, 0])
    for row in revenue: aggregate[str(row[10]).strip()][0] += row[15] if isinstance(row[15], (int, float)) else 0; aggregate[str(row[10]).strip()][1] += row[17] if isinstance(row[17], (int, float)) else 0
    for row in inventory: aggregate[str(row[2]).strip()][2] += row[11] if isinstance(row[11], (int, float)) else 0; aggregate[str(row[2]).strip()][3] += row[12] if isinstance(row[12], (int, float)) else 0
    for code, totals in aggregate.items():
        product_row = product.get(code, ()); product_sheet.append([code, product_row[2] if len(product_row) > 2 else "", product_row[6] if len(product_row) > 6 else "", product_row[21] if len(product_row) > 21 else "", product_row[22] if len(product_row) > 22 else "", *totals])
    def copy_source(sheet_name: str, key: str, source_sheet: str | None = None, min_row: int = 1) -> None:
        if key not in by: return
        source = load_workbook(io.BytesIO(by[key]), read_only=True, data_only=True); worksheet = source[source_sheet] if source_sheet and source_sheet in source.sheetnames else source.active; output = workbook.create_sheet(sheet_name)
        for row in worksheet.iter_rows(min_row=min_row, values_only=True): output.append(list(row))
    copy_source("Product detail", "product"); copy_source("Product final", "product"); copy_source("Purchase PO detail", "purchase", "LDL", 4)
    if "purchase" in by:
        source = load_workbook(io.BytesIO(by["purchase"]), read_only=True, data_only=True)["LDL"]; output = workbook.create_sheet("Purchase final"); output.append(list(next(source.iter_rows(min_row=4, max_row=4, values_only=True))))
        for row in purchase: output.append(list(row))
    copy_source("Inventory source", "inventory", None, 4)
    if "inventory" in by:
        source = load_workbook(io.BytesIO(by["inventory"]), read_only=True, data_only=True).active; output = workbook.create_sheet("Inventory final"); output.append(list(next(source.iter_rows(min_row=4, max_row=4, values_only=True))))
        for row in inventory: output.append(list(row))
    # Pre-order feedback is a reviewed register. Keep it visible in the final
    # workbook, but never treat its annotated rows as new mismatch candidates.
    copy_source("Pre-order source", "preorder"); copy_source("Pre-orders final", "preorder"); copy_source("Revenue raw", "revenue", None, 4)
    if "revenue" in by:
        source = load_workbook(io.BytesIO(by["revenue"]), read_only=True, data_only=True).active; output = workbook.create_sheet("Revenue final"); output.append(list(next(source.iter_rows(min_row=4, max_row=4, values_only=True))) + ["Net Revenue"])
        for row_no, row in enumerate(revenue, 2): output.append(list(row) + [f"=R{row_no}-MAX(0,U{row_no})-X{row_no}-Y{row_no}"])
    copy_source("CRM orders", "crm", "Danh sách"); copy_source("CRM items", "crm", "Bảng hàng hóa"); copy_source("Target detail", "target"); copy_source("Target final", "target")
    for sheet in workbook.worksheets:
        for cell in sheet[1]: cell.font = Font(color="FFFFFF", bold=True); cell.fill = PatternFill("solid", fgColor="1F4E78")
    output = io.BytesIO(); workbook.save(output)
    return PsiBuildResult(dict(rows), issues, gaps, output.getvalue())
