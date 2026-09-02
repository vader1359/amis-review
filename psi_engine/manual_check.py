from __future__ import annotations

from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from hashlib import sha256
import io
import json
from pathlib import Path
from typing import Any, BinaryIO, Iterable, Mapping

from openpyxl import load_workbook


ALLOWED_STATUSES = frozenset({"APPROVED", "OPEN", "REVIEW_REQUIRED", "RESOLVED", "EXPIRED"})
ALLOWED_ACTIONS = frozenset(
    {
        "KEEP AS MISMATCH",
        "IGNORE MISMATCH",
        "EXCLUDE FROM PREORDER",
        "EXCLUDE ORDER FROM PSI",
        "CLASSIFY FOC/COST-ONLY",
        "MAP SKU",
        "RESOLVED",
    }
)
ALLOWED_SCOPES = frozenset({"ORDER", "ORDER_SKU", "SKU", "CONTROL"})
ALLOWED_MATCH_TYPES = frozenset({"EXACT", "SUFFIX"})

EXCEPTION_HEADERS = (
    "Exception ID",
    "Fingerprint",
    "Status",
    "Action",
    "Scope",
    "Order ID",
    "Raw SKU",
    "Canonical SKU",
    "Quantity",
    "Value",
    "Issue Type",
    "Reason",
    "Evidence",
    "Effective From",
    "Effective To",
    "Approved By",
    "Approved Date",
    "First Seen",
    "Last Seen",
    "Notes",
)
PREORDER_HEADERS = (
    "Exclusion ID",
    "Fingerprint",
    "Status",
    "Action",
    "Order ID",
    "Raw SKU",
    "Canonical SKU",
    "Quantity",
    "Net Value",
    "Source Row",
    "Source No",
    "KT Note",
    "Treatment",
    "Effective From",
    "Effective To",
    "Approved By",
    "Approved Date",
    "Evidence",
    "Notes",
    "Disposition",
)
ORDER_EXCLUSION_HEADERS = (
    "Exclusion ID",
    "Fingerprint",
    "Status",
    "Action",
    "Scope",
    "Order ID",
    "Reason",
    "Treatment",
    "Effective From",
    "Effective To",
    "Approved By",
    "Approved Date",
    "Evidence",
    "Notes",
    "Disposition",
)
MAPPING_HEADERS = (
    "Mapping ID",
    "Fingerprint",
    "Status",
    "Action",
    "Match Type",
    "Source Scope",
    "Raw SKU / Pattern",
    "Canonical SKU / Replacement",
    "Reason",
    "Effective From",
    "Effective To",
    "Approved By",
    "Approved Date",
    "Evidence",
    "Notes",
)
REQUIRED_SHEETS = frozenset(
    {
        "Summary",
        "Exceptions",
        "Preorder Exclusions",
        "Order Exclusions",
        "SKU Mappings",
        "Source Archive",
        "Change Log",
        "Lists",
    }
)


class ManualCheckValidationError(ValueError):
    """Raised when Manual Check cannot be trusted by the deterministic pipeline."""


def _text(value: Any) -> str:
    return "" if value in (None, "") else str(value).strip()


def _upper(value: Any) -> str:
    return _text(value).upper()


def _number_text(value: Any) -> str:
    if value in (None, ""):
        return ""
    try:
        number = Decimal(str(value))
    except (InvalidOperation, ValueError):
        return _text(value)
    if number == number.to_integral():
        return str(number.quantize(Decimal("1")))
    return format(number.normalize(), "f")


def _decimal(value: Any) -> Decimal | None:
    if value in (None, ""):
        return None
    try:
        return Decimal(str(value))
    except (InvalidOperation, ValueError) as exc:
        raise ManualCheckValidationError(f"invalid numeric value: {value!r}") from exc


def _date(value: Any) -> date | None:
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = _text(value)
    try:
        return date.fromisoformat(text[:10])
    except ValueError as exc:
        raise ManualCheckValidationError(f"invalid ISO date: {value!r}") from exc


def make_fingerprint(
    *,
    action: str,
    scope: str,
    order_id: str = "",
    raw_sku: str = "",
    canonical_sku: str = "",
    quantity: Any = "",
    value: Any = "",
    issue_type: str = "",
) -> str:
    """Create the stable fingerprint used by both workbook and pipeline."""

    material = {
        "action": _upper(action),
        "canonical_sku": _upper(canonical_sku),
        "issue_type": _upper(issue_type),
        "order_id": _upper(order_id),
        "quantity": _number_text(quantity),
        "raw_sku": _upper(raw_sku),
        "scope": _upper(scope),
        "value": _number_text(value),
    }
    serialized = json.dumps(material, ensure_ascii=False, separators=(",", ":"), sort_keys=True)
    return sha256(serialized.encode("utf-8")).hexdigest()


def make_preorder_fingerprint(*, action: str, order_id: str, canonical_sku: str) -> str:
    """Stable permanent identity; Quantity and Net Value are audit snapshots."""
    material = {
        "action": _upper(action),
        "canonical_sku": _upper(canonical_sku),
        "order_id": _upper(order_id),
        "scope": "ORDER_SKU",
    }
    serialized = json.dumps(material, ensure_ascii=False, separators=(",", ":"), sort_keys=True)
    return sha256(serialized.encode("utf-8")).hexdigest()


def make_order_exclusion_fingerprint(*, action: str, scope: str, order_id: str) -> str:
    """Stable identity for an approved order-level PSI exclusion."""
    material = {
        "action": _upper(action),
        "order_id": _upper(order_id),
        "scope": _upper(scope),
    }
    serialized = json.dumps(material, ensure_ascii=False, separators=(",", ":"), sort_keys=True)
    return sha256(serialized.encode("utf-8")).hexdigest()


def _active(status: str, effective_from: date | None, effective_to: date | None, as_of: date) -> bool:
    return (
        status == "APPROVED"
        and (effective_from is None or effective_from <= as_of)
        and (effective_to is None or as_of <= effective_to)
    )


@dataclass(frozen=True, slots=True)
class ExceptionRule:
    exception_id: str
    fingerprint: str
    status: str
    action: str
    scope: str
    order_id: str
    raw_sku: str
    canonical_sku: str
    quantity: Decimal | None
    value: Decimal | None
    issue_type: str
    reason: str
    evidence: str
    effective_from: date | None
    effective_to: date | None
    approved_by: str
    approved_date: date | None
    notes: str

    def is_active(self, as_of: date) -> bool:
        return _active(self.status, self.effective_from, self.effective_to, as_of)


@dataclass(frozen=True, slots=True)
class PreorderExclusion:
    exclusion_id: str
    fingerprint: str
    status: str
    action: str
    order_id: str
    raw_sku: str
    canonical_sku: str
    quantity: Decimal
    net_value: Decimal
    source_row: int | None
    source_no: str
    kt_note: str
    treatment: str
    effective_from: date | None
    effective_to: date | None
    approved_by: str
    approved_date: date | None
    evidence: str
    notes: str
    disposition: str = ""

    def is_active(self, as_of: date) -> bool:
        return _active(self.status, self.effective_from, self.effective_to, as_of)


@dataclass(frozen=True, slots=True)
class OrderExclusion:
    exclusion_id: str
    fingerprint: str
    status: str
    action: str
    scope: str
    order_id: str
    reason: str
    treatment: str
    effective_from: date | None
    effective_to: date | None
    approved_by: str
    approved_date: date | None
    evidence: str
    notes: str
    disposition: str

    def is_active(self, as_of: date) -> bool:
        return _active(self.status, self.effective_from, self.effective_to, as_of)


@dataclass(frozen=True, slots=True)
class SkuMapping:
    mapping_id: str
    fingerprint: str
    status: str
    action: str
    match_type: str
    source_scope: str
    raw_sku_or_pattern: str
    canonical_sku_or_replacement: str
    reason: str
    effective_from: date | None
    effective_to: date | None
    approved_by: str
    approved_date: date | None
    evidence: str
    notes: str

    def is_active(self, as_of: date) -> bool:
        return _active(self.status, self.effective_from, self.effective_to, as_of)


@dataclass(frozen=True, slots=True)
class ManualCheckRegistry:
    exceptions: tuple[ExceptionRule, ...]
    preorder_exclusions: tuple[PreorderExclusion, ...]
    sku_mappings: tuple[SkuMapping, ...]
    order_exclusions: tuple[OrderExclusion, ...] = ()

    def summary(self, as_of: date | None = None) -> dict[str, int]:
        today = as_of or date.today()
        return {
            "exceptions": len(self.exceptions),
            "approved_active_exceptions": sum(rule.is_active(today) for rule in self.exceptions),
            "open_exceptions": sum(rule.status == "OPEN" for rule in self.exceptions),
            "approved_active_preorder_exclusions": sum(rule.is_active(today) for rule in self.preorder_exclusions),
            "approved_active_order_exclusions": sum(rule.is_active(today) for rule in self.order_exclusions),
            "approved_active_sku_mappings": sum(rule.is_active(today) for rule in self.sku_mappings),
        }

    def matching_order_exclusion(self, order_id: Any, as_of: date | None = None) -> OrderExclusion | None:
        today = as_of or date.today()
        order = _upper(order_id)
        for rule in self.order_exclusions:
            if (
                rule.is_active(today)
                and rule.action == "EXCLUDE ORDER FROM PSI"
                and rule.disposition == "PERMANENT / DONE"
                and order == rule.order_id
            ):
                return rule
        return None

    def map_sku(self, raw_sku: Any, source_scope: str = "ALL OFFICIAL SOURCES", as_of: date | None = None) -> str:
        current = _upper(raw_sku)
        today = as_of or date.today()
        scope = _upper(source_scope)
        active = [
            rule
            for rule in self.sku_mappings
            if rule.is_active(today)
            and rule.action == "MAP SKU"
            and (_upper(rule.source_scope) == "ALL OFFICIAL SOURCES" or _upper(rule.source_scope) == scope)
        ]
        for rule in active:
            if rule.match_type == "EXACT" and current == _upper(rule.raw_sku_or_pattern):
                return _upper(rule.canonical_sku_or_replacement)
        for rule in active:
            suffix = _upper(rule.raw_sku_or_pattern)
            if rule.match_type == "SUFFIX" and suffix and current.endswith(suffix):
                return current[: -len(suffix)] + _upper(rule.canonical_sku_or_replacement)
        return current

    def suppressing_exception(self, case: Mapping[str, Any], as_of: date | None = None) -> ExceptionRule | None:
        today = as_of or date.today()
        for rule in self.exceptions:
            if not rule.is_active(today) or rule.action != "IGNORE MISMATCH":
                continue
            if _exception_matches(rule, case, self, today):
                return rule
        return None

    def matching_preorder_exclusion(
        self,
        *,
        order_id: Any,
        sku: Any,
        quantity: Any,
        net_value: Any,
        as_of: date | None = None,
    ) -> PreorderExclusion | None:
        today = as_of or date.today()
        order = _upper(order_id)
        raw_sku = _upper(sku)
        canonical_sku = self.map_sku(raw_sku, as_of=today)
        for rule in self.preorder_exclusions:
            if not rule.is_active(today) or rule.action != "EXCLUDE FROM PREORDER":
                continue
            if (
                order == rule.order_id
                and canonical_sku == rule.canonical_sku
                and rule.disposition == "PERMANENT / DONE"
            ):
                return rule
        return None


def _exception_matches(rule: ExceptionRule, case: Mapping[str, Any], registry: ManualCheckRegistry, as_of: date) -> bool:
    order_id = _upper(case.get("order_id") or case.get("record_key"))
    raw_sku = _upper(case.get("raw_sku") or case.get("sku"))
    canonical_sku = _upper(case.get("canonical_sku")) or registry.map_sku(raw_sku, as_of=as_of)
    issue_type = _upper(case.get("issue_type"))
    quantity = _decimal(case.get("quantity"))
    value = _decimal(case.get("value"))
    if rule.scope == "ORDER":
        return order_id == rule.order_id and issue_type == _upper(rule.issue_type)
    if rule.scope == "ORDER_SKU":
        return (
            order_id == rule.order_id
            and canonical_sku == rule.canonical_sku
            and quantity == rule.quantity
            and value == rule.value
            and issue_type == _upper(rule.issue_type)
        )
    if rule.scope == "SKU":
        return canonical_sku == rule.canonical_sku and issue_type == _upper(rule.issue_type)
    return False


def partition_mismatches(
    cases: Iterable[Mapping[str, Any]],
    registry: ManualCheckRegistry,
    as_of: date | None = None,
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    """Return active cases and audit copies suppressed by exact approved rules."""

    active: list[dict[str, Any]] = []
    suppressed: list[dict[str, Any]] = []
    for original in cases:
        case = dict(original)
        rule = registry.suppressing_exception(case, as_of)
        if rule is None:
            active.append(case)
            continue
        suppressed.append(
            {
                **case,
                "status": "ignored",
                "suppressed": True,
                "manual_check_id": rule.exception_id,
                "manual_action": rule.action,
                "manual_reason": rule.reason,
                "manual_evidence": rule.evidence,
            }
        )
    return active, suppressed


def exclude_preorders(
    rows: Iterable[Mapping[str, Any]],
    registry: ManualCheckRegistry,
    as_of: date | None = None,
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    """Apply only approved exact-match preorder exclusions using semantic fields."""

    kept: list[dict[str, Any]] = []
    excluded: list[dict[str, Any]] = []
    for original in rows:
        row = dict(original)
        rule = registry.matching_preorder_exclusion(
            order_id=_first(row, "order_id", "Order ID", "ĐH", "Số đơn hàng"),
            sku=_first(row, "sku", "Raw SKU", "PRODUCT ID", "Mã hàng", "Mã hàng hóa"),
            quantity=_first(row, "quantity", "Quantity", "QUANTITY SOLD", "Số lượng"),
            net_value=_first(row, "net_value", "Net Value", "NET REV SOLD", "Giá trị"),
            as_of=as_of,
        )
        if rule is None:
            kept.append(row)
        else:
            excluded.append(
                {
                    **row,
                    "manual_check_id": rule.exclusion_id,
                    "manual_action": rule.action,
                    "manual_reason": rule.kt_note,
                    "manual_evidence": rule.evidence,
                }
            )
    return kept, excluded


def exclude_orders(
    rows: Iterable[Mapping[str, Any]],
    registry: ManualCheckRegistry,
    as_of: date | None = None,
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    """Remove approved cancelled orders from PSI business rows without mutating inputs."""
    kept: list[dict[str, Any]] = []
    excluded: list[dict[str, Any]] = []
    for original in rows:
        row = dict(original)
        rule = registry.matching_order_exclusion(
            _first(row, "order_id", "Order ID", "ĐH", "Số đơn hàng", "SALE ORDER"),
            as_of,
        )
        if rule is None:
            kept.append(row)
        else:
            excluded.append(
                {
                    **row,
                    "manual_check_id": rule.exclusion_id,
                    "manual_action": rule.action,
                    "manual_reason": rule.reason,
                    "manual_evidence": rule.evidence,
                }
            )
    return kept, excluded


def _first(row: Mapping[str, Any], *keys: str) -> Any:
    for key in keys:
        if key in row and row[key] not in (None, ""):
            return row[key]
    return ""


def load_manual_check(source: str | Path | bytes | bytearray | BinaryIO) -> ManualCheckRegistry:
    workbook_source: Any
    if isinstance(source, (bytes, bytearray)):
        workbook_source = io.BytesIO(bytes(source))
    else:
        workbook_source = source
    workbook = load_workbook(workbook_source, read_only=True, data_only=True)
    try:
        missing_sheets = sorted(REQUIRED_SHEETS - set(workbook.sheetnames))
        if missing_sheets:
            raise ManualCheckValidationError("Manual Check missing sheets: " + ", ".join(missing_sheets))
        exception_dicts = _sheet_records(workbook["Exceptions"], EXCEPTION_HEADERS)
        preorder_dicts = _sheet_records(workbook["Preorder Exclusions"], PREORDER_HEADERS)
        order_exclusion_dicts = _sheet_records(workbook["Order Exclusions"], ORDER_EXCLUSION_HEADERS)
        mapping_dicts = _sheet_records(workbook["SKU Mappings"], MAPPING_HEADERS)
        exceptions = tuple(_parse_exception(row) for row in exception_dicts)
        preorder_exclusions = tuple(_parse_preorder(row) for row in preorder_dicts)
        order_exclusions = tuple(_parse_order_exclusion(row) for row in order_exclusion_dicts)
        sku_mappings = tuple(_parse_mapping(row) for row in mapping_dicts)
        registry = ManualCheckRegistry(exceptions, preorder_exclusions, sku_mappings, order_exclusions)
        _validate_registry(registry)
        return registry
    finally:
        workbook.close()


def _sheet_records(worksheet: Any, expected_headers: tuple[str, ...]) -> list[dict[str, Any]]:
    expected = set(expected_headers)
    header_row = None
    headers: list[str] = []
    for row_number, row in enumerate(worksheet.iter_rows(min_row=1, max_row=10, values_only=True), start=1):
        candidate = [_text(value) for value in row]
        if expected.issubset(set(candidate)):
            header_row = row_number
            headers = candidate
            break
    if header_row is None:
        raise ManualCheckValidationError(f"{worksheet.title}: required semantic headers were not found")
    indexes = {header: headers.index(header) for header in expected_headers}
    records: list[dict[str, Any]] = []
    id_header = expected_headers[0]
    for row in worksheet.iter_rows(min_row=header_row + 1, values_only=True):
        record = {header: row[index] if index < len(row) else None for header, index in indexes.items()}
        if _text(record[id_header]):
            records.append(record)
    return records


def _parse_exception(row: Mapping[str, Any]) -> ExceptionRule:
    return ExceptionRule(
        exception_id=_text(row["Exception ID"]),
        fingerprint=_text(row["Fingerprint"]),
        status=_upper(row["Status"]),
        action=_upper(row["Action"]),
        scope=_upper(row["Scope"]),
        order_id=_upper(row["Order ID"]),
        raw_sku=_upper(row["Raw SKU"]),
        canonical_sku=_upper(row["Canonical SKU"]),
        quantity=_decimal(row["Quantity"]),
        value=_decimal(row["Value"]),
        issue_type=_text(row["Issue Type"]),
        reason=_text(row["Reason"]),
        evidence=_text(row["Evidence"]),
        effective_from=_date(row["Effective From"]),
        effective_to=_date(row["Effective To"]),
        approved_by=_text(row["Approved By"]),
        approved_date=_date(row["Approved Date"]),
        notes=_text(row["Notes"]),
    )


def _parse_preorder(row: Mapping[str, Any]) -> PreorderExclusion:
    return PreorderExclusion(
        exclusion_id=_text(row["Exclusion ID"]),
        fingerprint=_text(row["Fingerprint"]),
        status=_upper(row["Status"]),
        action=_upper(row["Action"]),
        order_id=_upper(row["Order ID"]),
        raw_sku=_upper(row["Raw SKU"]),
        canonical_sku=_upper(row["Canonical SKU"]),
        quantity=_decimal(row["Quantity"]) or Decimal("0"),
        net_value=_decimal(row["Net Value"]) or Decimal("0"),
        source_row=int(row["Source Row"]) if row["Source Row"] not in (None, "") else None,
        source_no=_text(row["Source No"]),
        kt_note=_text(row["KT Note"]),
        treatment=_text(row["Treatment"]),
        effective_from=_date(row["Effective From"]),
        effective_to=_date(row["Effective To"]),
        approved_by=_text(row["Approved By"]),
        approved_date=_date(row["Approved Date"]),
        evidence=_text(row["Evidence"]),
        notes=_text(row["Notes"]),
        disposition=_upper(row["Disposition"]),
    )


def _parse_order_exclusion(row: Mapping[str, Any]) -> OrderExclusion:
    return OrderExclusion(
        exclusion_id=_text(row["Exclusion ID"]),
        fingerprint=_text(row["Fingerprint"]),
        status=_upper(row["Status"]),
        action=_upper(row["Action"]),
        scope=_upper(row["Scope"]),
        order_id=_upper(row["Order ID"]),
        reason=_text(row["Reason"]),
        treatment=_text(row["Treatment"]),
        effective_from=_date(row["Effective From"]),
        effective_to=_date(row["Effective To"]),
        approved_by=_text(row["Approved By"]),
        approved_date=_date(row["Approved Date"]),
        evidence=_text(row["Evidence"]),
        notes=_text(row["Notes"]),
        disposition=_upper(row["Disposition"]),
    )


def _parse_mapping(row: Mapping[str, Any]) -> SkuMapping:
    return SkuMapping(
        mapping_id=_text(row["Mapping ID"]),
        fingerprint=_text(row["Fingerprint"]),
        status=_upper(row["Status"]),
        action=_upper(row["Action"]),
        match_type=_upper(row["Match Type"]),
        source_scope=_upper(row["Source Scope"]),
        raw_sku_or_pattern=_upper(row["Raw SKU / Pattern"]),
        canonical_sku_or_replacement=_upper(row["Canonical SKU / Replacement"]),
        reason=_text(row["Reason"]),
        effective_from=_date(row["Effective From"]),
        effective_to=_date(row["Effective To"]),
        approved_by=_text(row["Approved By"]),
        approved_date=_date(row["Approved Date"]),
        evidence=_text(row["Evidence"]),
        notes=_text(row["Notes"]),
    )


def _validate_registry(registry: ManualCheckRegistry) -> None:
    errors: list[str] = []
    _unique_ids("Exception ID", [rule.exception_id for rule in registry.exceptions], errors)
    _unique_ids("Exclusion ID", [rule.exclusion_id for rule in registry.preorder_exclusions], errors)
    _unique_ids("Order Exclusion ID", [rule.exclusion_id for rule in registry.order_exclusions], errors)
    _unique_ids("Mapping ID", [rule.mapping_id for rule in registry.sku_mappings], errors)

    for rule in registry.exceptions:
        _validate_common(rule.exception_id, rule.status, rule.action, rule.effective_from, rule.effective_to, rule.approved_by, rule.approved_date, errors)
        if rule.scope not in ALLOWED_SCOPES:
            errors.append(f"{rule.exception_id}: invalid scope {rule.scope!r}")
        if rule.status == "OPEN" and rule.action != "KEEP AS MISMATCH":
            errors.append(f"{rule.exception_id}: OPEN must use KEEP AS MISMATCH")
        if rule.status == "APPROVED" and rule.action == "IGNORE MISMATCH" and (not rule.order_id or not rule.issue_type):
            errors.append(f"{rule.exception_id}: approved mismatch suppression requires exact Order ID + Issue Type")
        expected = make_fingerprint(
            action=rule.action,
            scope=rule.scope,
            order_id=rule.order_id,
            raw_sku=rule.raw_sku,
            canonical_sku=rule.canonical_sku,
            quantity=rule.quantity,
            value=rule.value,
            issue_type=rule.issue_type,
        )
        if rule.fingerprint != expected:
            errors.append(f"{rule.exception_id}: fingerprint does not match row content")

    for rule in registry.preorder_exclusions:
        _validate_common(rule.exclusion_id, rule.status, rule.action, rule.effective_from, rule.effective_to, rule.approved_by, rule.approved_date, errors)
        if rule.action != "EXCLUDE FROM PREORDER":
            errors.append(f"{rule.exclusion_id}: Preorder Exclusions action must be EXCLUDE FROM PREORDER")
        if not rule.order_id or not rule.raw_sku or not rule.canonical_sku:
            errors.append(f"{rule.exclusion_id}: exact Order ID + Raw SKU + Canonical SKU are required")
        if rule.status == "APPROVED" and rule.disposition != "PERMANENT / DONE":
            errors.append(f"{rule.exclusion_id}: approved Pre-order exclusion must be PERMANENT / DONE")
        expected = make_preorder_fingerprint(
            action=rule.action,
            order_id=rule.order_id,
            canonical_sku=rule.canonical_sku,
        )
        if rule.fingerprint != expected:
            errors.append(f"{rule.exclusion_id}: fingerprint does not match row content")

    for rule in registry.order_exclusions:
        _validate_common(rule.exclusion_id, rule.status, rule.action, rule.effective_from, rule.effective_to, rule.approved_by, rule.approved_date, errors)
        if rule.action != "EXCLUDE ORDER FROM PSI":
            errors.append(f"{rule.exclusion_id}: Order Exclusions action must be EXCLUDE ORDER FROM PSI")
        if rule.scope != "ALL PSI BUSINESS SHEETS":
            errors.append(f"{rule.exclusion_id}: Order Exclusions scope must be ALL PSI BUSINESS SHEETS")
        if not rule.order_id:
            errors.append(f"{rule.exclusion_id}: Order ID is required")
        if not rule.reason or not rule.evidence:
            errors.append(f"{rule.exclusion_id}: reason and evidence are required")
        if rule.status == "APPROVED" and rule.disposition != "PERMANENT / DONE":
            errors.append(f"{rule.exclusion_id}: approved order exclusion must be PERMANENT / DONE")
        expected = make_order_exclusion_fingerprint(
            action=rule.action,
            scope=rule.scope,
            order_id=rule.order_id,
        )
        if rule.fingerprint != expected:
            errors.append(f"{rule.exclusion_id}: fingerprint does not match row content")

    for rule in registry.sku_mappings:
        _validate_common(rule.mapping_id, rule.status, rule.action, rule.effective_from, rule.effective_to, rule.approved_by, rule.approved_date, errors)
        if rule.action != "MAP SKU":
            errors.append(f"{rule.mapping_id}: SKU Mappings action must be MAP SKU")
        if rule.match_type not in ALLOWED_MATCH_TYPES:
            errors.append(f"{rule.mapping_id}: invalid match type {rule.match_type!r}")
        if not rule.raw_sku_or_pattern or not rule.canonical_sku_or_replacement:
            errors.append(f"{rule.mapping_id}: raw and canonical mapping values are required")
        expected = make_fingerprint(
            action=rule.action,
            scope="SKU",
            raw_sku=rule.raw_sku_or_pattern,
            canonical_sku=rule.canonical_sku_or_replacement,
            issue_type=f"{rule.match_type}:{rule.source_scope}",
        )
        if rule.fingerprint != expected:
            errors.append(f"{rule.mapping_id}: fingerprint does not match row content")

    if errors:
        raise ManualCheckValidationError("Manual Check validation failed:\n- " + "\n- ".join(errors[:50]))


def _validate_common(
    record_id: str,
    status: str,
    action: str,
    effective_from: date | None,
    effective_to: date | None,
    approved_by: str,
    approved_date: date | None,
    errors: list[str],
) -> None:
    if status not in ALLOWED_STATUSES:
        errors.append(f"{record_id}: invalid status {status!r}")
    if action not in ALLOWED_ACTIONS:
        errors.append(f"{record_id}: invalid action {action!r}")
    if effective_from and effective_to and effective_from > effective_to:
        errors.append(f"{record_id}: Effective From is after Effective To")
    if status == "APPROVED" and (not approved_by or approved_date is None):
        errors.append(f"{record_id}: APPROVED requires Approved By + Approved Date")


def _unique_ids(label: str, values: list[str], errors: list[str]) -> None:
    seen: set[str] = set()
    for value in values:
        if not value:
            errors.append(f"{label}: blank ID")
        elif value in seen:
            errors.append(f"{label}: duplicate {value}")
        seen.add(value)
