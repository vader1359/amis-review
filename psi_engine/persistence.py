from dataclasses import asdict, dataclass, replace
from datetime import UTC, date, datetime, timedelta
from hashlib import sha256
import io
import json
import os
from pathlib import PurePosixPath
from typing import Final, Protocol
from urllib.parse import urlencode
from urllib.request import Request, urlopen
from uuid import uuid4

from openpyxl import load_workbook

from .engine import build, classify
from .sources import OPTIONAL_SOURCES, REQUIRED_SOURCES, UPLOAD_SOURCES

MAX_UPLOAD_BYTES: Final = 25 * 1024 * 1024
ALLOWED_CONTENT_TYPE: Final = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
TABLES: Final = ("upload_batches", "source_snapshots", "source_snapshot_metadata", "source_selections", "reconciliation_runs", "reconciliation_run_sources", "normalized_records", "mismatches", "mismatch_history", "known_issues", "psi_drafts", "draft_sources", "psi_releases", "release_sources", "activity_logs")


def week_to_period(week: str) -> tuple[str, str]:
    try:
        year, week_number = week.split("-W")
        end = date.fromisocalendar(int(year), int(week_number), 7)
    except (ValueError, TypeError):
        raise UploadValidationError("week must use ISO format YYYY-Www") from None
    return week, end.isoformat()


class UploadAuthorizationError(PermissionError):
    pass


class UploadValidationError(ValueError):
    pass


class PsiRepository(Protocol):
    def insert(self, table: str, row: dict[str, str | int | list[str] | dict[str, str] | None], auth_token: str = "") -> None: ...
    def insert_many(self, table: str, rows: list[dict[str, object]], auth_token: str = "") -> None: ...
    def upload(self, path: str, content: bytes, auth_token: str = "") -> None: ...
    def lookup(self, table: str, filters: dict[str, str], auth_token: str = "") -> list[dict[str, str | int | list[str] | dict[str, str] | None]]: ...
    def upsert(self, table: str, row: dict[str, str | int | list[str] | dict[str, str] | None], conflict: str, auth_token: str = "") -> None: ...
    def transition_mismatch(self, mismatch_id: str, to_status: str, comment: str, evidence: dict[str, str], actor_id: str, auth_token: str = "") -> None: ...
    def download(self, path: str, auth_token: str = "") -> bytes: ...
    def signed_download(self, path: str, expires: int, auth_token: str = "") -> str: ...
    def delete(self, table: str, filters: dict[str, str], auth_token: str = "") -> None: ...


class PsiMemoryRepository:
    def __init__(self) -> None:
        self.data: dict[str, list[dict[str, str | int | list[str] | dict[str, str] | None]]] = {table: [] for table in TABLES}
        self.objects: dict[str, bytes] = {}

    def insert(self, table: str, row: dict[str, str | int | list[str] | dict[str, str] | None], auth_token: str = "") -> None:
        self.data[table].append(row.copy())

    def insert_many(self, table: str, rows: list[dict[str, object]], auth_token: str = "") -> None:
        self.data[table].extend(row.copy() for row in rows)

    def upsert(self, table: str, row: dict[str, str | int | list[str] | dict[str, str] | None], conflict: str, auth_token: str = "") -> None:
        matches = [stored for stored in self.data[table] if str(stored.get(conflict)) == str(row.get(conflict))]
        if matches:
            matches[0].update(row)
        else:
            self.insert(table, row)

    def upload(self, path: str, content: bytes, auth_token: str = "") -> None:
        if path in self.objects:
            raise UploadValidationError("immutable object path already exists")
        self.objects[path] = content

    def download(self, path: str, auth_token: str = "") -> bytes:
        if path not in self.objects:
            raise UploadValidationError("release object is unavailable")
        return self.objects[path]

    def delete(self, table: str, filters: dict[str, str], auth_token: str = "") -> None:
        self.data[table] = [row for row in self.data[table] if not all(str(row.get(key)) == value for key, value in filters.items())]

    def signed_download(self, path: str, expires: int, auth_token: str = "") -> str:
        raise UploadValidationError("local repository does not issue Supabase signed URLs")

    def lookup(self, table: str, filters: dict[str, str], auth_token: str = "") -> list[dict[str, str | int | list[str] | dict[str, str] | None]]:
        return [row for row in self.data[table] if all(str(row.get(key)) == value for key, value in filters.items())]

    def rows(self, table: str) -> list[dict[str, str | int | list[str] | dict[str, str] | None]]:
        return list(self.data[table])

    def resolve_mismatches(self, run_id: str, actor_id: str) -> None:
        for mismatch in self.data["mismatches"]:
            if mismatch["reconciliation_run_id"] == run_id:
                previous = str(mismatch["status"])
                mismatch["status"] = "resolved"
                self.insert("mismatch_history", {"mismatch_id": str(mismatch["id"]), "changed_by": actor_id, "from_status": previous, "to_status": "resolved", "comment": None, "evidence": {}})

    def transition_mismatch(self, mismatch_id: str, to_status: str, comment: str, evidence: dict[str, str], actor_id: str, auth_token: str = "") -> None:
        matches = self.lookup("mismatches", {"id": mismatch_id})
        if not matches:
            raise UploadValidationError("mismatch is unavailable")
        mismatch = matches[0]
        previous = str(mismatch["status"])
        allowed = {"new": {"assigned"}, "assigned": {"in_progress"}, "in_progress": {"resolved", "known", "ignored"}, "resolved": {"reopened"}, "known": {"reopened"}, "ignored": {"reopened"}, "reopened": {"assigned", "in_progress"}}
        if to_status not in allowed or to_status not in allowed.get(previous, set()):
            raise UploadValidationError(f"invalid mismatch transition: {previous} -> {to_status}")
        mismatch["status"] = to_status
        self.insert("mismatch_history", {"mismatch_id": mismatch_id, "changed_by": actor_id, "from_status": previous, "to_status": to_status, "comment": comment, "evidence": evidence})
        if to_status in {"resolved", "known", "ignored"}:
            fingerprint = str(mismatch.get("fingerprint", ""))
            if fingerprint:
                self.upsert(
                    "known_issues",
                    {
                        "fingerprint": fingerprint,
                        "title": str(mismatch.get("record_key") or fingerprint),
                        "reason": json.dumps({"action": to_status, "comment": comment, "evidence": evidence}, ensure_ascii=False, separators=(",", ":")),
                        "status": "approved" if to_status == "resolved" else "known",
                        "created_by": actor_id,
                    },
                    "fingerprint",
                )
        self.insert("activity_logs", {"actor_id": actor_id, "action": "mismatch.transitioned", "entity_type": "mismatch", "entity_id": mismatch_id, "metadata": {"to_status": to_status}})


class SupabaseRepository:
    def __init__(self, url: str, service_role_key: str, bucket: str = "psi-source") -> None:
        if not url or not service_role_key:
            raise UploadValidationError("Supabase server configuration is incomplete")
        self.url, self.key, self.bucket = url.rstrip("/"), service_role_key, bucket

    @classmethod
    def from_environment(cls) -> "SupabaseRepository":
        return cls(os.environ.get("SUPABASE_URL", ""), os.environ.get("SUPABASE_SERVICE_ROLE_KEY", ""))

    def _user_bearer(self, auth_token: str) -> str:
        bearer = auth_token.strip()
        if not bearer:
            raise UploadAuthorizationError("authenticated bearer is required")
        return bearer

    def _service_bearer(self) -> str:
        return self.key

    def insert(self, table: str, row: dict[str, str | int | list[str] | dict[str, str] | None], auth_token: str = "") -> None:
        bearer = self._service_bearer()
        request = Request(f"{self.url}/rest/v1/{table}", data=json.dumps(row).encode(), method="POST", headers={"apikey": self.key, "Authorization": f"Bearer {bearer}", "Content-Type": "application/json", "Prefer": "return=minimal"})
        with urlopen(request, timeout=10):
            pass

    def insert_many(self, table: str, rows: list[dict[str, object]], auth_token: str = "") -> None:
        bearer = self._service_bearer()
        for offset in range(0, len(rows), 500):
            request = Request(f"{self.url}/rest/v1/{table}", data=json.dumps(rows[offset:offset + 500]).encode(), method="POST", headers={"apikey": self.key, "Authorization": f"Bearer {bearer}", "Content-Type": "application/json", "Prefer": "return=minimal"})
            with urlopen(request, timeout=60):
                pass

    def insert_activity_log(self, row: dict[str, str | int | list[str] | dict[str, str] | None], auth_token: str = "") -> None:
        self.insert("activity_logs", row, auth_token)

    def upsert(self, table: str, row: dict[str, str | int | list[str] | dict[str, str] | None], conflict: str, auth_token: str = "") -> None:
        bearer = self._service_bearer()
        request = Request(f"{self.url}/rest/v1/{table}?on_conflict={conflict}", data=json.dumps(row).encode(), method="POST", headers={"apikey": self.key, "Authorization": f"Bearer {bearer}", "Content-Type": "application/json", "Prefer": "resolution=merge-duplicates,return=minimal"})
        with urlopen(request, timeout=10):
            pass

    def lookup(self, table: str, filters: dict[str, str], auth_token: str = "") -> list[dict[str, str | int | list[str] | dict[str, str] | None]]:
        query = urlencode({key: f"eq.{value}" for key, value in filters.items()})
        bearer = self._service_bearer()
        request = Request(f"{self.url}/rest/v1/{table}?{query}", method="GET", headers={"apikey": self.key, "Authorization": f"Bearer {bearer}"})
        with urlopen(request, timeout=10) as response:
            payload = json.loads(response.read().decode("utf-8"))
        if not isinstance(payload, list):
            raise UploadValidationError("Supabase lookup returned an invalid response")
        return [row for row in payload if isinstance(row, dict)]

    def delete(self, table: str, filters: dict[str, str], auth_token: str = "") -> None:
        query = urlencode({key: f"eq.{value}" for key, value in filters.items()})
        bearer = self._service_bearer()
        request = Request(f"{self.url}/rest/v1/{table}?{query}", method="DELETE", headers={"apikey": self.key, "Authorization": f"Bearer {bearer}"})
        with urlopen(request, timeout=10):
            pass

    def authenticated_actor(self, auth_token: str) -> str:
        bearer = self._user_bearer(auth_token)
        request = Request(f"{self.url}/auth/v1/user", method="GET", headers={"apikey": self.key, "Authorization": f"Bearer {bearer}"})
        with urlopen(request, timeout=10) as response:
            payload = json.loads(response.read().decode("utf-8"))
        if not isinstance(payload, dict) or not isinstance(payload.get("id"), str):
            raise UploadAuthorizationError("authenticated actor is unavailable")
        return payload["id"]

    def resolve_context(self, team_id: str, actor_id: str, period_key: str, source_type: str, auth_token: str) -> dict[str, str]:
        authenticated_actor = self.authenticated_actor(auth_token)
        if authenticated_actor != actor_id or not self.lookup("profiles", {"id": authenticated_actor}, auth_token):
            raise UploadAuthorizationError("authenticated actor is unavailable")
        profile = self.lookup("profiles", {"id": authenticated_actor}, auth_token)
        membership = self.lookup("team_memberships", {"team_id": team_id, "profile_id": authenticated_actor}, auth_token)
        period = self.lookup("reporting_periods", {"period_key": period_key}, auth_token)
        rule = self.lookup("rule_versions", {"version": "1"}, auth_token)
        if not profile or not membership or not period or not rule:
            raise UploadAuthorizationError("Supabase actor, team membership, period, or rule context is unavailable")
        period_id = str(period[0]["id"])
        rule_id = str(rule[0]["id"])
        snapshots = self.lookup("source_snapshots", {"team_id": team_id, "reporting_period_id": period_id, "source_type": source_type}, auth_token)
        return {"reporting_period_id": period_id, "rule_version_id": rule_id, "version": str(len(snapshots) + 1)}

    def upload(self, path: str, content: bytes, auth_token: str = "") -> None:
        bearer = self._service_bearer()
        request = Request(f"{self.url}/storage/v1/object/{self.bucket}/{path}", data=content, method="POST", headers={"apikey": self.key, "Authorization": f"Bearer {bearer}", "Content-Type": ALLOWED_CONTENT_TYPE})
        with urlopen(request, timeout=10):
            pass

    def download(self, path: str, auth_token: str = "") -> bytes:
        bearer = self._service_bearer()
        request = Request(f"{self.url}/storage/v1/object/{self.bucket}/{path}", method="GET", headers={"apikey": self.key, "Authorization": f"Bearer {bearer}"})
        with urlopen(request, timeout=10) as response:
            return response.read()

    def signed_download(self, path: str, expires: int, auth_token: str = "") -> str:
        bearer = self._service_bearer()
        request = Request(f"{self.url}/storage/v1/object/sign/{self.bucket}/{path}", data=json.dumps({"expiresIn": expires}).encode(), method="POST", headers={"apikey": self.key, "Authorization": f"Bearer {bearer}", "Content-Type": "application/json"})
        with urlopen(request, timeout=10) as response:
            payload = json.loads(response.read().decode("utf-8"))
        if not isinstance(payload, dict) or not isinstance(payload.get("signedURL"), str):
            raise UploadValidationError("Supabase signed URL response is invalid")
        return payload["signedURL"]

    def transition_mismatch(self, mismatch_id: str, to_status: str, comment: str, evidence: dict[str, str], actor_id: str, auth_token: str = "") -> None:
        bearer = self._user_bearer(auth_token)
        if to_status not in {"resolved", "known", "ignored"}:
            raise UploadValidationError("final mismatch status is required")
        if self.authenticated_actor(bearer) != actor_id:
            raise UploadAuthorizationError("authenticated actor is unavailable")
        rows = self.lookup("mismatches", {"id": mismatch_id}, auth_token)
        if not rows:
            raise UploadValidationError("mismatch is unavailable")
        mismatch = rows[0]
        fingerprint = str(mismatch.get("fingerprint", ""))
        if not fingerprint:
            raise UploadValidationError("mismatch fingerprint is unavailable")
        reason = json.dumps({"action": to_status, "comment": comment, "evidence": evidence}, ensure_ascii=False, separators=(",", ":"))
        self.upsert(
            "known_issues",
            {
                "fingerprint": fingerprint,
                "title": str(mismatch.get("record_key") or fingerprint),
                "reason": reason,
                "status": "approved" if to_status == "resolved" else "known",
                "created_by": actor_id,
            },
            "fingerprint",
            auth_token,
        )


@dataclass(frozen=True, slots=True)
class UploadRequest:
    team_id: str; actor_id: str; reporting_period: str; data_as_of: str; source_type: str; filename: str; content: bytes; content_type: str = ALLOWED_CONTENT_TYPE; rule_version_id: str = ""; reporting_period_id: str = ""; version: int = 0


@dataclass(frozen=True, slots=True)
class Snapshot:
    id: str; team_id: str; reporting_period: str; source_type: str; version: int; original_filename: str; object_path: str; checksum_sha256: str; byte_size: int; data_as_of: str; schema_status: str; header_preview: tuple[str, ...]; schema_gaps: tuple[tuple[str | int | float | None, ...], ...]; row_count: int; uploaded_by: str; uploaded_at: str


@dataclass(frozen=True, slots=True)
class RunProvenance:
    id: str; reporting_period: str; source_snapshot_ids: tuple[str, ...]; normalized_material: tuple[tuple[str | int | float | None, ...], ...]; fingerprint: str; duplicate_policy: str; status: str; started_by: str; started_at: str


@dataclass(frozen=True, slots=True)
class PersistedUpload:
    snapshot: Snapshot; run: RunProvenance; draft_id: str

    def to_json(self) -> dict[str, object]:
        return {"snapshot": asdict(self.snapshot), "run": asdict(self.run), "draft_id": self.draft_id}


class PsiMemoryStore:
    def __init__(self, repository: PsiRepository | None = None) -> None:
        self.repository = repository or PsiMemoryRepository()
        self.snapshots: list[Snapshot] = []; self.runs: list[RunProvenance] = []; self.activity: list[dict[str, str]] = []; self.selections: dict[tuple[str, ...], str] = {}; self.drafts: list[str] = []; self.actor_teams: dict[str, set[str]] = {"user-a": {"team-a"}}

    def persist(self, request: UploadRequest, auth_token: str = "") -> PersistedUpload:
        if "-W" in request.reporting_period:
            week, week_end = week_to_period(request.reporting_period)
            request = replace(request, reporting_period=week, data_as_of=request.data_as_of or week_end)
        if isinstance(self.repository, SupabaseRepository):
            context = self.repository.resolve_context(request.team_id, request.actor_id, request.reporting_period, request.source_type, auth_token)
            request = replace(request, reporting_period_id=context["reporting_period_id"], rule_version_id=context["rule_version_id"], version=int(context["version"]))
        if not isinstance(self.repository, SupabaseRepository):
            self._authorize(request)
        self._validate_metadata(request)
        source_type = classify(request.filename)
        if source_type != request.source_type: raise UploadValidationError("filename does not match source_type")
        checksum = sha256(request.content).hexdigest(); version = request.version or sum(s.team_id == request.team_id and s.reporting_period == request.reporting_period and s.source_type == request.source_type for s in self.snapshots) + 1
        snapshot_id = str(uuid4()); batch_id = str(uuid4()); now = datetime.now(UTC).isoformat(); headers, row_count = workbook_metadata(request.content, request.source_type)
        result = build(self._reconciliation_files(request, auth_token)); gaps = tuple(tuple(cell for cell in gap) for gap in result.gaps); snapshot = Snapshot(snapshot_id, request.team_id, request.reporting_period, request.source_type, version, request.filename, f"{request.team_id}/{batch_id}/{snapshot_id}", checksum, len(request.content), request.data_as_of, "failed" if gaps else "passed", headers, gaps, row_count, request.actor_id, now)
        material = tuple(tuple(cell for cell in issue) for issue in result.issues + result.gaps); fingerprint = sha256((request.reporting_period + request.source_type + json_material(material)).encode()).hexdigest(); run = RunProvenance(str(uuid4()), request.reporting_period, (snapshot_id,), material, fingerprint, "active_mismatches_recorded", "completed", request.actor_id, now); draft_id = str(uuid4())
        self._write(snapshot, run, draft_id, request, result.issues, auth_token); return PersistedUpload(snapshot, run, draft_id)

    def _reconciliation_files(self, request: UploadRequest, auth_token: str) -> dict[str, bytes]:
        files = {request.filename: request.content}
        related = {"product": {"purchase", "revenue"}, "purchase": {"product"}, "revenue": {"product"}}.get(request.source_type, set())
        if not related:
            return files
        if isinstance(self.repository, SupabaseRepository):
            selections = self.repository.lookup("source_selections", {"reporting_period_id": request.reporting_period_id}, auth_token)
            for selection in selections:
                snapshots = self.repository.lookup("source_snapshots", {"id": str(selection["source_snapshot_id"])}, auth_token)
                if not snapshots or str(snapshots[0]["source_type"]) not in related:
                    continue
                snapshot = snapshots[0]
                files[str(snapshot["original_filename"])] = self.repository.download(str(snapshot["object_path"]), auth_token)
            return files
        latest: dict[str, Snapshot] = {}
        for snapshot in self.snapshots:
            if snapshot.team_id == request.team_id and snapshot.reporting_period == request.reporting_period and snapshot.source_type in related:
                if snapshot.source_type not in latest or snapshot.version > latest[snapshot.source_type].version:
                    latest[snapshot.source_type] = snapshot
        for snapshot in latest.values():
            files[snapshot.original_filename] = self.repository.download(snapshot.object_path, auth_token)
        return files

    def _write(self, snapshot: Snapshot, run: RunProvenance, draft_id: str, request: UploadRequest, issues: list[list[str | int | float | None]], auth_token: str) -> None:
        period_id = request.reporting_period_id or request.reporting_period; batch_id = snapshot.object_path.split("/")[1]
        self.repository.insert("upload_batches", {"id": batch_id, "team_id": request.team_id, "reporting_period_id": period_id, "uploaded_by": request.actor_id}, auth_token); self.repository.insert("source_snapshots", {"id": snapshot.id, "upload_batch_id": batch_id, "team_id": snapshot.team_id, "reporting_period_id": period_id, "source_type": snapshot.source_type, "version": snapshot.version, "original_filename": snapshot.original_filename, "object_path": snapshot.object_path, "checksum_sha256": snapshot.checksum_sha256, "byte_size": snapshot.byte_size, "data_as_of": snapshot.data_as_of, "schema_status": snapshot.schema_status, "row_count": snapshot.row_count, "uploaded_by": snapshot.uploaded_by}, auth_token); self.repository.insert("source_snapshot_metadata", {"source_snapshot_id": snapshot.id, "header_preview": list(snapshot.header_preview), "schema_gaps": [list(gap) for gap in snapshot.schema_gaps]}, auth_token); self.repository.upsert("source_selections", {"reporting_period_id": period_id, "source_type": request.source_type, "source_snapshot_id": snapshot.id, "selected_by": request.actor_id}, "reporting_period_id,source_type", auth_token); self.repository.insert("reconciliation_runs", {"id": run.id, "reporting_period_id": period_id, "started_by": request.actor_id, "rule_version_id": request.rule_version_id, "status": run.status, "started_at": run.started_at, "completed_at": run.started_at}, auth_token); self.repository.insert("reconciliation_run_sources", {"reconciliation_run_id": run.id, "source_snapshot_id": snapshot.id, "source_type": request.source_type}, auth_token); self.repository.insert("normalized_records", {"id": str(uuid4()), "reconciliation_run_id": run.id, "source_snapshot_id": snapshot.id, "source_type": request.source_type, "record_key": run.fingerprint, "normalized_values": {"material": json_material(run.normalized_material)}, "source_row_number": 1}, auth_token); self.repository.upload(snapshot.object_path, request.content, auth_token)
        existing_fingerprints = {str(row.get("fingerprint")) for row in self.repository.lookup("mismatches", {"reporting_period_id": period_id}, auth_token)}
        existing_fingerprints.update(
            str(row.get("fingerprint"))
            for row in self.repository.lookup("known_issues", {}, auth_token)
            if str(row.get("status")) in {"known", "approved"}
        )
        mismatch_rows: list[dict[str, object]] = []
        source_types = {"Revenue": "revenue", "Purchase/PO": "purchase"}
        for issue in issues:
            if len(issue) < 7:
                continue
            source, filename, sheet, row_number, code, description, message = issue[:7]
            source_type = source_types.get(str(source), request.source_type)
            identity = json.dumps({"source_type": source_type, "record_key": code or description, "issue": message}, ensure_ascii=False, separators=(",", ":"), sort_keys=True)
            issue_fingerprint = sha256(identity.encode()).hexdigest()
            if issue_fingerprint in existing_fingerprints:
                continue
            existing_fingerprints.add(issue_fingerprint)
            mismatch_rows.append({"id": str(uuid4()), "reconciliation_run_id": run.id, "reporting_period_id": period_id, "rule_version_id": request.rule_version_id, "source_type": source_type, "record_key": str(code or description or issue_fingerprint), "fingerprint": issue_fingerprint, "severity": "blocking", "status": "new", "values_by_source": {"file": str(filename), "sheet": str(sheet), "row": row_number, "code": str(code or ""), "description": str(description or ""), "issue": str(message)}})
        self.repository.insert_many("mismatches", mismatch_rows, auth_token)
        self.repository.insert("psi_drafts", {"id": draft_id, "reporting_period_id": period_id, "reconciliation_run_id": run.id, "rule_version_id": request.rule_version_id, "status": "pending_review", "created_by": request.actor_id}, auth_token); self.repository.insert("draft_sources", {"draft_id": draft_id, "source_snapshot_id": snapshot.id, "source_type": request.source_type}, auth_token); activity = {"team_id": request.team_id, "actor_id": request.actor_id, "action": "psi.upload.persisted", "entity_type": "source_snapshot", "entity_id": snapshot.id, "metadata": {"run_id": run.id}}; self.repository.insert_activity_log(activity, auth_token) if isinstance(self.repository, SupabaseRepository) else self.repository.insert("activity_logs", activity, auth_token); self.snapshots.append(snapshot); self.runs.append(run); key = (request.team_id, request.reporting_period, request.source_type) if "-W" in request.reporting_period else (request.reporting_period, request.source_type); self.selections[key] = snapshot.id; self.drafts.append(draft_id); self.activity.append({"action": "psi.upload.persisted", "actor_id": request.actor_id, "team_id": request.team_id})

    def _authorize(self, request: UploadRequest) -> None:
        if request.actor_id == "anonymous-uploader":
            return
        if request.team_id not in self.actor_teams.get(request.actor_id, set()): raise UploadAuthorizationError("actor is not a member of team")

    def weekly_status(self, team_id: str, week: str) -> dict[str, object]:
        week_to_period(week)
        latest: dict[str, Snapshot] = {}
        for snapshot in self.snapshots:
            if snapshot.team_id == team_id and snapshot.reporting_period == week and snapshot.source_type in UPLOAD_SOURCES:
                if snapshot.source_type not in latest or snapshot.version > latest[snapshot.source_type].version:
                    latest[snapshot.source_type] = snapshot
        files = {source: {"status": "uploaded", "version": latest[source].version, "snapshot_id": latest[source].id, "filename": latest[source].original_filename} if source in latest else {"status": "missing", "version": None, "snapshot_id": None, "filename": None} for source in UPLOAD_SOURCES}
        payload: dict[str, object] = {
            "team_id": team_id,
            "week": week,
            "files": files,
            "required_sources": list(REQUIRED_SOURCES),
            "optional_sources": list(OPTIONAL_SOURCES),
            "owned_sources": list(UPLOAD_SOURCES),
            "ready": all(source in latest for source in REQUIRED_SOURCES),
        }
        if payload["ready"]:
            from psi_engine.release import PsiReleaseService, ReleaseRequest
            record = PsiReleaseService(self).generate(ReleaseRequest(week, "anonymous-uploader", team_id))
            payload["download_url"] = record.signed_url
        return payload

    @staticmethod
    def _validate_metadata(request: UploadRequest) -> None:
        if not request.content or len(request.content) > MAX_UPLOAD_BYTES: raise UploadValidationError("upload byte size is outside the allowed range")
        if request.content_type != ALLOWED_CONTENT_TYPE or not request.filename.lower().endswith(".xlsx"): raise UploadValidationError("content type and extension must be xlsx")
        if PurePosixPath(request.filename).name != request.filename or not request.filename.strip(): raise UploadValidationError("filename must be a single safe path component")
        try:
            if "-W" in request.reporting_period:
                week_to_period(request.reporting_period)
            else:
                datetime.strptime(request.reporting_period, "%Y-%m")
            datetime.strptime(request.data_as_of, "%Y-%m-%d")
        except ValueError as error:
            raise UploadValidationError("reporting period and data_as_of must be real calendar dates") from error
        if "-W" not in request.reporting_period and request.reporting_period != request.data_as_of[:7]:
            raise UploadValidationError("data_as_of must be in reporting period")


def workbook_metadata(content: bytes, source_type: str) -> tuple[tuple[str, ...], int]:
    workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True); worksheet = workbook["LDL"] if source_type == "purchase" and "LDL" in workbook.sheetnames else workbook.active; header_row = 4 if source_type in {"purchase", "revenue"} else 1; rows = list(worksheet.iter_rows(min_row=header_row, values_only=True)); header = tuple(str(value).strip() for value in rows[0] if value not in (None, "")); return header, max(0, sum(any(value not in (None, "") for value in row) for row in rows[1:]))


def json_material(issues: tuple[tuple[str | int | float | None, ...], ...]) -> str:
    return json.dumps(issues, ensure_ascii=False, separators=(",", ":"), sort_keys=True)
