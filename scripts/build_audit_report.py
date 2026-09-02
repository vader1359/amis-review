#!/usr/bin/env python3
"""
AMIS CRM ↔ MISA Comprehensive Audit Report Generator
Cross-references all 5 input files, detects discrepancies, outputs Excel report.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from collections import defaultdict, Counter
from datetime import datetime, date
from pathlib import Path
import csv, re, sys, textwrap

ROOT = Path(__file__).resolve().parents[1]
BASE = ROOT / 'input'
OLD_CHECK = ROOT / 'old_check'
TODAY = date(2026, 7, 5)  # Snapshot date

# ============================================================
# 1. LOAD DATA
# ============================================================

print("[1/5] Loading CRM_Sale.xlsx...", file=sys.stderr)

crm_orders = {}  # order_id -> dict
crm_lineitems = defaultdict(list)  # order_id -> list of line items

wb = openpyxl.load_workbook(f'{BASE}/CRM_Sale.xlsx', read_only=True, data_only=True)

ws = wb['Danh sách']
crm_headers = [str(c.value) if c.value is not None else f'COL{i}' for i, c in enumerate(ws[1])]
# Build header index
crm_h = {h: i for i, h in enumerate(crm_headers)}

row_count = 0
for row in ws.iter_rows(min_row=2, values_only=True):
    if row[0] is None:
        continue
    oid = str(row[0]).strip()
    if oid.startswith('CKHO'):
        continue
    rec = {}
    for hdr, idx in crm_h.items():
        rec[hdr] = row[idx]
    crm_orders[oid] = rec
    row_count += 1

print(f"   CRM Danh sách: {row_count} orders loaded", file=sys.stderr)

# Load line items
ws2 = wb['Bảng hàng hóa']
li_headers = [str(c.value) if c.value is not None else f'COL{i}' for i, c in enumerate(ws2[1])]
li_h = {h: i for i, h in enumerate(li_headers)}
li_count = 0
for row in ws2.iter_rows(min_row=2, values_only=True):
    if row[0] is None:
        continue
    oid = str(row[0]).strip()
    if oid.startswith('CKHO'):
        continue
    rec = {}
    for hdr, idx in li_h.items():
        rec[hdr] = row[idx]
    crm_lineitems[oid].append(rec)
    li_count += 1
print(f"   CRM Bảng hàng hóa: {li_count} line items loaded", file=sys.stderr)
wb.close()

# ============================================================

print("[2/5] Loading MISA_Accounting.xlsx...", file=sys.stderr)

misa_orders = {}
wb = openpyxl.load_workbook(f'{BASE}/MISA_Accounting.xlsx', read_only=True, data_only=True)
ws = wb['Đơn đặt hàng']
misa_headers = [str(c.value) if c.value is not None else f'COL{i}' for i, c in enumerate(ws[3])]
misa_h = {h: i for i, h in enumerate(misa_headers)}
misa_count = 0
for row in ws.iter_rows(min_row=4, values_only=True):
    if row[misa_h['Số đơn hàng']] is None:
        continue
    oid = str(row[misa_h['Số đơn hàng']]).strip()
    if oid.startswith('CKHO'):
        continue
    rec = {}
    for hdr, idx in misa_h.items():
        rec[hdr] = row[idx]
    misa_orders[oid] = rec
    misa_count += 1
print(f"   MISA: {misa_count} orders loaded", file=sys.stderr)
wb.close()

# ============================================================

print("[3/5] Loading So_chi_tiet_ban_hang.xlsx...", file=sys.stderr)

so_orders = defaultdict(list)  # order_id -> list of invoice lines
so_invoice_map = defaultdict(set)  # order_id -> set of invoice numbers

wb = openpyxl.load_workbook(f'{BASE}/So_chi_tiet_ban_hang.xlsx', read_only=True, data_only=True)
ws = wb['SỔ CHI TIẾT BÁN HÀNG']
so_headers = [str(c.value) if c.value is not None else f'COL{i}' for i, c in enumerate(ws[4])]
so_h = {h: i for i, h in enumerate(so_headers)}
so_row_count = 0
dh_in_so_col = set()
for row in ws.iter_rows(min_row=5, values_only=True):
    so_row_count += 1
    # Try col 29 first
    oid = str(row[28]).strip() if row[28] else ''
    desc = str(row[8]).strip() if row[8] else ''
    
    if not oid:
        # Try extracting from description: "Đơn hàng bán DH-xxxx"
        m = re.search(r'DH-\d+', desc)
        if m:
            oid = m.group(0)
    
    if oid and oid.startswith('DH-'):
        dh_in_so_col.add(oid)
        rec = {}
        for hdr, idx in so_h.items():
            rec[hdr] = row[idx]
        so_orders[oid].append(rec)
        inv_num = str(row[7]).strip() if row[7] else ''
        if inv_num:
            so_invoice_map[oid].add(inv_num)

print(f"   So_chi_tiet: {so_row_count} rows, {len(dh_in_so_col)} unique DH- orders", file=sys.stderr)
print(f"   Orders with invoices: {len(so_invoice_map)}", file=sys.stderr)
wb.close()

# ============================================================

print("[4/5] Loading Pre order feedback...", file=sys.stderr)

pre_orders = {}
pre_lines = defaultdict(list)
wb = openpyxl.load_workbook(f'{BASE}/Pre order feedback.xlsx', read_only=True, data_only=True)
ws = wb['Pre-orders']
pre_headers = [str(c.value) if c.value is not None else f'COL{i}' for i, c in enumerate(ws[1])]
pre_h = {h: i for i, h in enumerate(pre_headers)}

for row in ws.iter_rows(min_row=2, values_only=True):
    if row[pre_h['ĐH']] is None:
        continue
    oid = str(row[pre_h['ĐH']]).strip()
    if oid.startswith('CKHO'):
        continue
    note = str(row[pre_h['Note']]).strip() if row[pre_h['Note']] else ''
    hangiao = str(row[pre_h['HẠN GIAO HÀNG']])[:10] if row[pre_h['HẠN GIAO HÀNG']] else ''
    
    if oid not in pre_orders:
        pre_orders[oid] = {
            'note_actions': set(),
            'is_pending': False,
            'is_resolved': False,
            'hangiao': hangiao,
            'total_rev': 0,
            'sales_date': row[pre_h['SALES DATE']],
        }
    if note:
        pre_orders[oid]['note_actions'].add(note)
    pre_orders[oid]['hangiao'] = hangiao
    if row[pre_h['NET REV SOLD']]:
        pre_orders[oid]['total_rev'] += float(row[pre_h['NET REV SOLD']])

for oid, info in pre_orders.items():
    has_pending = any('Pre order' in n for n in info['note_actions'])
    has_revenue = any('ghi nhận' in n for n in info['note_actions'])
    has_action = any(n not in ('Pre order', 'Đã ghi nhận ở revenue', '') for n in info['note_actions'])
    info['is_pending'] = has_pending and not has_revenue
    info['is_resolved'] = has_revenue and not has_pending and not has_action
    info['needs_action'] = has_action

print(f"   Pre-order: {len(pre_orders)} unique orders", file=sys.stderr)
wb.close()

# ============================================================

print("[5/5] Loading Master audit CSV...", file=sys.stderr)

master_orders = {}
with open(f'{OLD_CHECK}/danh_sach_toan_bo_don_hang_review_nhom_loi_huong_xu_ly.csv', 'r', encoding='utf-8-sig') as f:
    reader = csv.DictReader(f)
    for row in reader:
        rec = {k.lower().strip(): v.strip() if v else '' for k, v in row.items() if k}
        oid = rec.get('order_id', '').strip()
        if not oid or oid.startswith('CKHO'):
            continue
        master_orders[oid] = rec

print(f"   Master: {len(master_orders)} orders loaded", file=sys.stderr)

# ============================================================
# 1b. LOAD CRM_Activities
# ============================================================

print("[1b] Loading CRM_Activities.xlsx...", file=sys.stderr)

activity_orders = defaultdict(list)  # order_id -> list of activities
activity_types = Counter()

wb = openpyxl.load_workbook(f'{BASE}/CRM_Activities.xlsx', read_only=True, data_only=True)
ws = wb['Danh sách']
act_headers = [str(c.value) if c.value is not None else f'COL{i}' for i, c in enumerate(ws[1])]
act_h = {h: i for i, h in enumerate(act_headers)}

act_count = 0
for row in ws.iter_rows(min_row=2, values_only=True):
    # Liên quan đến = col 6
    ref = str(row[6]).strip() if row[6] else ''
    if not ref:
        continue
    # Extract DH-xxx from "Liên quan đến" field
    oids = re.findall(r'DH-\d+', ref)
    if not oids:
        continue
    act_type = str(row[0]).strip() if row[0] else ''
    activity_types[act_type] += 1
    
    rec = {}
    for hdr, idx in act_h.items():
        rec[hdr] = row[idx]
    
    for oid in oids:
        activity_orders[oid].append(rec)
    act_count += 1

print(f"   Activities: {act_count} loaded, types: {dict(activity_types.most_common())}", file=sys.stderr)
wb.close()


# ============================================================
# 2. DETECTION RULES
# ============================================================

print("\nRunning detection rules...", file=sys.stderr)

discrepancies = []
disc_id = [0]  # mutable counter

def add_disc(discrepancy_type, severity, order_id, crm_status, misa_status, detail, 
             source_files='', suggestion='', old_group='', priority=''):
    master = master_orders.get(order_id, {})
    old_group_val = old_group or master.get('primary_issue_group', '')
    old_comment = master.get('review_comment', '')
    old_suggested_fix = master.get('suggested_fix', '')

    disc_id[0] += 1
    disc = {
        'id': disc_id[0],
        'type': discrepancy_type,
        'severity': severity,
        'order_id': order_id,
        'customer': crm_orders.get(order_id, {}).get('Khách hàng', ''),
        'owner': crm_orders.get(order_id, {}).get('Người thực hiện', ''),
        'gioi_doan': crm_orders.get(order_id, {}).get('Giai đoạn', ''),
        'CRM_delivery': str(crm_status.get('delivery', '')),
        'CRM_payment': str(crm_status.get('payment', '')),
        'CRM_accounting': str(crm_status.get('accounting', '')),
        'CRM_approval': str(crm_status.get('approval', '')),
        'CRM_execution': str(crm_status.get('execution', '')),
        'CRM_invoiced': str(crm_status.get('invoiced', '')),
        'CRM_invoice_value': str(crm_status.get('invoice_value', '')),
        'MISA_delivery': str(misa_status.get('delivery', '')),
        'MISA_payment': str(misa_status.get('payment', '')),
        'MISA_accounting': str(misa_status.get('accounting', '')),
        'MISA_invoiced': str(misa_status.get('invoiced', '')),
        'MISA_invoice_value': str(misa_status.get('invoice_value', '')),
        'MISA_thuc_thu': str(misa_status.get('thuc_thu', '')),
        'MISA_con_thu': str(misa_status.get('con_thu', '')),
        'MISA_delivery_date': str(misa_status.get('delivery_date', '')),
        'detail': detail,
        'source_files': source_files,
        'suggestion': suggestion,
        'old_group': old_group_val,
        'old_comment': old_comment,
        'old_suggested_fix': old_suggested_fix,
        'owner_chinh': '',
    }
    discrepancies.append(disc)

# Common helper
def val(v):
    """Safely convert value to string"""
    if v is None:
        return ''
    return str(v).strip()

def float_val(v):
    """Convert to float, return 0 if None/empty"""
    if v is None:
        return 0.0
    try:
        return float(v)
    except (ValueError, TypeError):
        return 0.0

# ============================================================

def to_date(raw):
    if raw is None:
        return None
    if isinstance(raw, (datetime, date)):
        if isinstance(raw, datetime):
            return raw.date()
        return raw
    text = val(raw)
    if not text:
        return None
    for fmt in ('%Y-%m-%d', '%d/%m/%Y'):
        try:
            return datetime.strptime(text[:10], fmt).date()
        except ValueError:
            pass
    return None

# Collect ALL order IDs across all sources
all_order_ids = set(crm_orders.keys()) | set(misa_orders.keys()) | set(so_orders.keys()) | set(pre_orders.keys()) | set(master_orders.keys())
print(f"Total unique orders across all sources before filtering: {len(all_order_ids)}", file=sys.stderr)

# Filter out orders before 31/12/2023
filtered_order_ids = set()
cutoff = date(2023, 12, 31)
for oid in all_order_ids:
    order_date = None
    if oid in crm_orders and crm_orders[oid].get('Ngày tạo'):
        order_date = to_date(crm_orders[oid].get('Ngày tạo'))
    if not order_date and oid in misa_orders and misa_orders[oid].get('Ngày đơn hàng'):
        order_date = to_date(misa_orders[oid].get('Ngày đơn hàng'))
    if not order_date and oid in pre_orders and pre_orders[oid].get('sales_date'):
        order_date = to_date(pre_orders[oid].get('sales_date'))
    if not order_date and oid in so_orders:
        dates = []
        for r in so_orders[oid]:
            d = to_date(r.get('Ngày chứng từ')) or to_date(r.get('Ngày hạch toán'))
            if d:
                dates.append(d)
        if dates:
            order_date = min(dates)
    if not order_date and oid in master_orders and master_orders[oid].get('order_date'):
        order_date = to_date(master_orders[oid].get('order_date'))

    if order_date and order_date < cutoff:
        continue
    filtered_order_ids.add(oid)

all_order_ids = filtered_order_ids
print(f"Total unique orders after excluding before 31/12/2023: {len(all_order_ids)}", file=sys.stderr)

# Status values
DELIVERY_DELIVERED = {'Đã giao hàng', 'Đã giao', 'Đã giao đủ'}
DELIVERY_DELIVERING = {'Đang giao hàng'}
DELIVERY_NOT_DELIVERED = {'Chưa giao hàng', ''}
PAID = {'Đã thanh toán', 'Đã thanh toán một phần'}
NOT_PAID = {'Chưa thanh toán', ''}
REJECTED = {'Từ chối ghi'}
DRAFT = {'Bản nháp'}
SUBMITTED = {'Đề nghị ghi'}
APPROVED = {'Đã duyệt'}

def parse_iso_date(raw):
    """Parse YYYY-MM-DD-like strings from Excel values."""
    text = val(raw)
    if not text:
        return None
    try:
        return datetime.strptime(text[:10], '%Y-%m-%d').date()
    except ValueError:
        return None

def check_processed_status(oid):
    """Note gift and voucher orders as 'HỢP LÝ HD' for processed status."""
    crm = crm_orders.get(oid, {})
    if not crm:
        return ""
    crm_ctkm = val(crm.get('CTKM áp dụng', '')).lower()
    crm_note = val(crm.get('Ghi chú đơn hàng', '')).lower()
    crm_gia_tri = float_val(crm.get('Giá trị đơn hàng', 0))
    
    # Check if this order is a gift or voucher
    is_gift = (
        crm_gia_tri == 0 or
        'gift' in crm_ctkm or
        'gift' in crm_note or
        'quà tặng' in crm_note or
        'tặng' in crm_note or
        'quà' in crm_note or
        'voucher' in crm_ctkm or
        'voucher' in crm_note
    )
    if is_gift:
        return "HỢP LÝ HD"
    return ""

def is_reportable_crm_order(accounting_status, approval_status):
    """Orders worth reporting: only approved orders on CRM."""
    return approval_status == 'Đã duyệt'

def norm_accounting(status):
    s = val(status)
    if s in ('Đã ghi', 'Đã ghi doanh số'):
        return 'Đã ghi'
    if s in ('Từ chối ghi', 'Từ chối ghi doanh số'):
        return 'Từ chối ghi'
    if s in ('Đề nghị ghi', 'Chưa ghi doanh số'):
        return 'Chưa ghi'
    if s in ('Bản nháp',):
        return 'Bản nháp'
    return s

loop_count = 0
for oid in sorted(all_order_ids):
    loop_count += 1
    if loop_count % 1000 == 0:
        print(f"   Processing: {loop_count}/{len(all_order_ids)}", file=sys.stderr)
    
    crm = crm_orders.get(oid, {})
    misa = misa_orders.get(oid, {})
    pre = pre_orders.get(oid, {})
    master = master_orders.get(oid, {})
    activities = activity_orders.get(oid, [])
    
    crm_stage = val(crm.get('Giai đoạn', ''))
    
    # CRM status values
    crm_delivery = val(crm.get('Tình trạng giao hàng', ''))
    crm_payment = val(crm.get('Tình trạng thanh toán', ''))
    crm_accounting = val(crm.get('Tình trạng ghi doanh số', ''))
    crm_approval = val(crm.get('Trạng thái phê duyệt', ''))
    crm_execution = val(crm.get('Tình trạng', ''))
    crm_invoiced = val(crm.get('Đã xuất hóa đơn', ''))
    crm_inv_val = float_val(crm.get('Giá trị đã xuất hóa đơn', ''))
    crm_con_thu = float_val(crm.get('Còn phải thu', ''))
    crm_gia_tri = float_val(crm.get('Giá trị đơn hàng', ''))
    
    # MISA status values
    misa_delivery = val(misa.get('Tình trạng giao hàng', ''))
    misa_payment_thucthu = float_val(misa.get('Thực thu', ''))
    misa_con_thu = float_val(misa.get('Số còn phải thu', ''))
    misa_accounting = val(misa.get('Tình trạng ghi doanh số', ''))
    misa_invoiced = val(misa.get('Tình trạng xuất hóa đơn', ''))
    misa_inv_val = float_val(misa.get('Giá trị đã xuất hóa đơn', ''))
    misa_delivery_date = val(misa.get('Ngày giao hàng', ''))
    misa_hangiao = val(misa.get('Hạn giao hàng', ''))
    misa_status = val(misa.get('Tình trạng', ''))
    misa_gia_tri = float_val(misa.get('Giá trị đơn hàng', ''))
    is_reportable = oid not in crm_orders or is_reportable_crm_order(crm_accounting, crm_approval)
    
    # So_chi_tiet data
    so_lines = so_orders.get(oid, [])
    so_invoices = so_invoice_map.get(oid, set())
    so_total = sum(float_val(r.get('Doanh số bán', 0)) for r in so_lines)
    so_has_invoice = len(so_invoices) > 0 or len(so_lines) > 0
    
    # Activities
    has_nvgh = any('Xác nhận giao hàng' in val(a.get('Loại nhiệm vụ', '')) for a in activities)
    has_payment_confirm = any('Xác nhận thanh toán' in val(a.get('Loại nhiệm vụ', '')) for a in activities)
    
    # Old master classification
    old_group = master.get('primary_issue_group', '')
    old_priority = master.get('priority', '')
    old_comment = master.get('review_comment', '')
    
    # CRM status struct for add_disc
    crm_status = {
        'delivery': crm_delivery,
        'payment': crm_payment,
        'accounting': crm_accounting,
        'approval': crm_approval,
        'execution': crm_execution,
        'invoiced': crm_invoiced,
        'invoice_value': crm_inv_val,
    }
    misa_status = {
        'delivery': misa_delivery,
        'payment': '' if misa_con_thu == 0 and misa_payment_thucthu == 0 else f'Thực thu={misa_payment_thucthu:,.0f}',
        'accounting': misa_accounting,
        'invoiced': misa_invoiced,
        'invoice_value': misa_inv_val,
        'thuc_thu': f'{misa_payment_thucthu:,.0f}' if misa_payment_thucthu else '0',
        'con_thu': f'{misa_con_thu:,.0f}',
        'delivery_date': misa_delivery_date,
    }
    
    # --- RULE A: CRM-only orders (not in MISA at all) ---
    if oid not in misa_orders and is_reportable:
        extra_comment = " (Có vẻ là lỗi)" if crm_accounting == 'Bản nháp' else ""
        add_disc('A-Missing_MISA', 'HIGH', oid, crm_status, {'delivery': 'NOT_IN_MISA'},
                 f"Đơn có trên CRM (ghi DS='{crm_accounting}', duyệt='{crm_approval}') nhưng KHÔNG có trên MISA{extra_comment}",
                 source_files='CRM_Sale',
                 suggestion='Kiểm tra đồng bộ CRM→MISA. Có thể đơn chưa được duyệt hoặc chưa sync.',
                 old_group=old_group)
    
    # --- RULE A2: MISA-only orders (not in CRM) ---
    # Skip CKHO-* (chứng từ kho, internal warehouse, not sales orders)
    if oid.startswith('CKHO'):
        pass
    elif oid not in crm_orders:
        add_disc('A2-Missing_CRM', 'HIGH', oid, {'delivery': 'NOT_IN_CRM'}, misa_status,
                 f"Đơn có trên MISA nhưng KHÔNG có trên CRM",
                 source_files='MISA_Accounting',
                 suggestion='Kiểm tra đồng bộ MISA→CRM. Có thể xóa trên CRM nhưng MISA còn.')
    
    # --- RULE B: Delivery status mismatch ---
    if is_reportable and crm_delivery in DELIVERY_DELIVERED and misa_delivery not in DELIVERY_DELIVERED and misa_delivery:
        extra_comment = " (Có vẻ lỗi, note lại)" if misa_delivery == 'Đang giao' else " (Kiểm lại có thể lỗi)" if misa_delivery == 'Chưa giao' else ""
        add_disc('B-Delivery_Mismatch', 'HIGH', oid, crm_status, misa_status,
                 f"CRM: '{crm_delivery}' nhưng MISA: '{misa_delivery}'{extra_comment}",
                 source_files='CRM_Sale+MISA_Accounting',
                 suggestion='Kiểm tra trạng thái giao hàng thực tế. Nếu MISA chưa giao, CRM đang sai.',
                 old_group=old_group)
    
    if is_reportable and crm_delivery not in DELIVERY_DELIVERED and misa_delivery in DELIVERY_DELIVERED and misa_delivery_date:
        add_disc('B2-Delivery_Mismatch', 'MEDIUM', oid, crm_status, misa_status,
                 f"MISA đã giao ({misa_delivery_date}) nhưng CRM: '{crm_delivery}'",
                 source_files='CRM_Sale+MISA_Accounting',
                 suggestion='CRM chưa cập nhật trạng thái giao hàng. Cần sales update.',
                 old_group=old_group)
    
    # --- RULE C: Payment mismatch ---
    if is_reportable and crm_payment in PAID and misa_payment_thucthu == 0 and misa_gia_tri > 0:
        add_disc('C-Payment_Mismatch', 'HIGH', oid, crm_status, misa_status,
                 f"CRM: '{crm_payment}' (giá trị đơn hàng={crm_gia_tri:,.0f}), MISA Thực thu=0 (có thể là đơn hàng quà tặng)",
                 source_files='CRM_Sale+MISA_Accounting',
                 suggestion='CRM nói đã thanh toán nhưng MISA không thấy tiền. Kiểm tra đơn hàng quà tặng hoặc sai sót.',
                 old_group=old_group)
    
    if is_reportable and crm_payment in NOT_PAID and misa_payment_thucthu > 0 and crm_gia_tri > 0:
        add_disc('C2-Payment_Mismatch', 'MEDIUM', oid, crm_status, misa_status,
                 f"CRM: '{crm_payment}' nhưng MISA Thực thu={misa_payment_thucthu:,.0f}",
                 source_files='CRM_Sale+MISA_Accounting',
                 suggestion='CRM chưa cập nhật thanh toán. Kiểm tra và update.',
                 old_group=old_group)
    
    # --- RULE D: Accounting record mismatch ---
    # Skip if CRM is still in draft (not yet submitted for processing)
    if not is_reportable:
        pass
    elif crm_accounting and misa_accounting and norm_accounting(crm_accounting) != norm_accounting(misa_accounting):
        add_disc('D-Accounting_Mismatch', 'HIGH', oid, crm_status, misa_status,
                 f"CRM: '{crm_accounting}' vs MISA: '{misa_accounting}'",
                 source_files='CRM_Sale+MISA_Accounting',
                 suggestion='Lệch trạng thái ghi doanh số. Cần đồng bộ.',
                 old_group=old_group)
    
    # --- RULE E: Invoice mismatch ---
    crm_invoice_date = val(crm.get('Ngày hóa đơn', ''))
    crm_invoice_no = val(crm.get('Số hóa đơn', ''))
    crm_has_invoice_info = crm_invoiced or (crm_inv_val > 0) or crm_invoice_date or crm_invoice_no

    if is_reportable and crm_has_invoice_info and not so_has_invoice:
        crm_inv_details = []
        if crm_inv_val > 0:
            crm_inv_details.append(f"giá trị HĐ={crm_inv_val:,.0f}")
        if crm_invoice_date:
            crm_inv_details.append(f"ngày xuất HĐ={crm_invoice_date}")
        if crm_invoice_no:
            crm_inv_details.append(f"số HĐ={crm_invoice_no}")
        crm_inv_str = f" ({', '.join(crm_inv_details)})" if crm_inv_details else f" ('{crm_invoiced}')"
        add_disc('E-Invoice_Mismatch', 'HIGH', oid, crm_status, misa_status,
                 f"CRM ghi nhận xuất hóa đơn{crm_inv_str} nhưng không có trong Sổ chi tiết bán hàng",
                 source_files='CRM_Sale+So_chi_tiet_ban_hang',
                 suggestion='CRM đã xuất hóa đơn nhưng không tìm thấy trong MISA Sổ chi tiết. Kiểm tra chứng từ.',
                 old_group=old_group)

    if is_reportable and so_has_invoice and not crm_has_invoice_info:
        crm_inv_details = []
        if crm_inv_val > 0:
            crm_inv_details.append(f"giá trị HĐ={crm_inv_val:,.0f}")
        if crm_invoice_date:
            crm_inv_details.append(f"ngày xuất HĐ={crm_invoice_date}")
        if crm_invoice_no:
            crm_inv_details.append(f"số HĐ={crm_invoice_no}")
        crm_inv_str = f" (CRM có: {', '.join(crm_inv_details)})" if crm_inv_details else " (CRM trống thông tin HĐ)"

        add_disc('E2-Invoice_Mismatch', 'MEDIUM', oid, crm_status, misa_status,
                 f"Có trong Sổ chi tiết bán hàng ({len(so_lines)} dòng, invoices={so_invoices}) nhưng CRM không ghi nhận xuất hóa đơn{crm_inv_str}",
                 source_files='CRM_Sale+So_chi_tiet_ban_hang',
                 suggestion='MISA đã có hóa đơn nhưng CRM chưa cập nhật. Cần đồng bộ.',
                 old_group=old_group)
    
    # --- RULE F: Rejected orders ---
    # Business decision: rejected accounting orders are excluded from the active issue report.
    
    # --- RULE G: Overdue delivery ---
    hg_date = parse_iso_date(misa_hangiao)
    if is_reportable and hg_date and hg_date < TODAY and crm_delivery not in DELIVERY_DELIVERED:
        add_disc('G-Overdue', 'MEDIUM', oid, crm_status, misa_status,
                 f"Quá hạn giao: Hạn={misa_hangiao[:10]}, quá {(TODAY - hg_date).days} ngày, vẫn '{crm_delivery}'",
                 source_files='MISA_Accounting+CRM_Sale',
                 suggestion='Đơn quá hạn giao hàng. Cần kiểm tra, điều chỉnh lại hạn giao hàng hoặc trạng thái giao hàng.',
                 old_group=old_group)
    
    # --- RULE H: Pre-order specific ---
    if pre:
        if pre.get('needs_action'):
            action_notes = [n for n in pre['note_actions'] if n not in ('Pre order', 'Đã ghi nhận ở revenue', '')]
            detail = '; '.join(action_notes[:3])
            if len(action_notes) > 3:
                detail += f' (+{len(action_notes)-3} more)'
            add_disc('H-Preorder_Action', 'MEDIUM', oid, crm_status, misa_status,
                     f"Pre-order cần xử lý: {detail}",
                     source_files='Pre order feedback',
                     suggestion='Xem chi tiết trong sheet Pre-order Issues.',
                     old_group=old_group)
    
    # --- RULE I: NVGH but delivery still open (G1 check) ---
    if is_reportable and has_nvgh and crm_delivery not in DELIVERY_DELIVERED:
        extra_comment = ""
        if crm_delivery == '':
            extra_comment = " NVGH"
        elif crm_delivery == 'Đang giao hàng':
            extra_comment = " (Có giao hết hàng chưa?)"
        elif crm_delivery == 'Chưa giao hàng':
            extra_comment = " (Kiểm lại, có thể lỗi)"
        add_disc('I-NVGH_Activity_Open_Delivery', 'LOW', oid, crm_status, misa_status,
                 f"Có activity NVGH nhưng CRM delivery vẫn '{crm_delivery}'{extra_comment}",
                 source_files='CRM_Activities+CRM_Sale',
                 suggestion='NVGH chỉ là activity xác nhận, không tự đóng delivery. Kiểm tra thực tế rồi cập nhật trạng thái giao hàng nếu đã giao.',
                 old_group=old_group if old_group else 'G1')
    
    # --- RULE J: Payment confirmed but still open (G2 check) ---
    if is_reportable and crm_payment in PAID and crm_delivery not in DELIVERY_DELIVERED and hg_date and hg_date < TODAY:
        old_g2 = (old_group == 'G2')
        if old_g2 or not old_group:
            add_disc('J-Paid_Not_Delivered', 'MEDIUM', oid, crm_status, misa_status,
                 f"Đã thanh toán ({crm_payment}) nhưng quá hẹn giao {misa_hangiao[:10]} và delivery='{crm_delivery}'",
                 source_files='CRM_Sale',
                 suggestion='Đã thanh toán nhưng quá hạn giao. Kiểm tra đã giao thực tế chưa, điền lại hạn giao hàng hoặc update delivery.',
                 old_group=old_group if old_group else 'G2')
    
    # --- RULE K: Old master audit unresolved check ---
    if is_reportable and old_group:
        # Check if old issues are still present
        if old_group == 'G1' and crm_delivery not in DELIVERY_DELIVERED:
            add_disc('K-Old_G1_Unresolved', 'HIGH', oid, crm_status, misa_status,
                     f"[Master cũ] G1: Có NVGH từ audit cũ nhưng delivery vẫn '{crm_delivery}'",
                     source_files='Master+CRM_Sale+CRM_Activities',
                     suggestion='Audit cũ đã phát hiện, vẫn chưa fix. Cần KT+Ops xử lý.',
                     old_group='G1')
        elif old_group == 'G2' and crm_delivery not in DELIVERY_DELIVERED:
            add_disc('K-Old_G2_Unresolved', 'HIGH', oid, crm_status, misa_status,
                     f"[Master cũ] G2: Đã TT chưa giao từ audit cũ, delivery vẫn '{crm_delivery}'",
                     source_files='Master+CRM_Sale',
                     suggestion='Audit cũ đã phát hiện, vẫn chưa fix. Cần sales update.',
                     old_group='G2')
        elif old_group == 'G7' and crm_accounting not in REJECTED:
            add_disc('K-Old_G7_MaybeFixed', 'LOW', oid, crm_status, misa_status,
                     f"[Master cũ] G7: Từng bị hủy/từ chối, nay CRM accounting='{crm_accounting}' — kiểm tra lại",
                     source_files='Master+CRM_Sale',
                     suggestion='Có thể đã được fix. Xác nhận và đóng issue.',
                     old_group='G7')
    
    # --- RULE L: Draft orders ---
    # Business decision: draft orders are excluded from this reconciliation report.
    
    # --- RULE M: CRM invoice value vs MISA invoice value ---
    if is_reportable and crm_inv_val > 0 and misa_inv_val > 0:
        diff = abs(crm_inv_val - misa_inv_val)
        if diff > 1000 and (diff / max(crm_inv_val, 0.01)) > 0.05:  # >5% diff
            add_disc('M-Invoice_Value_Mismatch', 'HIGH', oid, crm_status, misa_status,
                     f"Lệch giá trị xuất HĐ: CRM={crm_inv_val:,.0f} vs MISA={misa_inv_val:,.0f} (diff={diff:,.0f})",
                     source_files='CRM_Sale+MISA_Accounting',
                     suggestion='Giá trị hóa đơn lệch. Kiểm tra chứng từ gốc.',
                     old_group=old_group)

    # --- RULE N: No activities at all ---
    if is_reportable and len(activities) == 0 and oid in crm_orders:
        created_str = val(crm.get('Ngày tạo', ''))[:10]
        add_disc('N-No_Activity', 'LOW', oid, crm_status, misa_status,
                 f"Không có hoạt động nào (từ {created_str})",
                 source_files='CRM_Activities+CRM_Sale',
                 suggestion='Đơn không có activity. Kiểm tra đơn ma hoặc đã abandoned.',
                 old_group=old_group)

print(f"\nTotal discrepancies found: {len(discrepancies)}", file=sys.stderr)

# Compute stats
severity_counts = Counter(d['severity'] for d in discrepancies)
type_counts = Counter(d['type'] for d in discrepancies)
print(f"By severity: {dict(severity_counts)}", file=sys.stderr)
print(f"By type: {dict(type_counts.most_common())}", file=sys.stderr)

# ============================================================
# 3. GENERATE EXCEL REPORT
# ============================================================

print("\nGenerating Excel report...", file=sys.stderr)

wb_out = openpyxl.Workbook()

# --- Styles ---
header_font = Font(bold=True, color='FFFFFF', size=11)
header_fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
cell_align = Alignment(vertical='top', wrap_text=False)

red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
yellow_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
orange_fill = PatternFill(start_color='F4B183', end_color='F4B183', fill_type='solid')

thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

def write_sheet(ws, title, headers, rows, col_widths=None):
    """Write a data sheet with headers and rows."""
    # Headers
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
    
    # Data
    for r, row in enumerate(rows, 2):
        for c, val in enumerate(row, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.alignment = cell_align
            cell.border = thin_border
    
    # Auto-width (not too wide)
    if col_widths:
        for c, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(c)].width = w
    else:
        for c, h in enumerate(headers, 1):
            ws.column_dimensions[get_column_letter(c)].width = min(max(len(str(h)) + 2, 15), 40)
    
    # Auto-filter
    ws.auto_filter.ref = f'A1:{get_column_letter(len(headers))}{len(rows)+1}'
    
    # Freeze top row
    ws.freeze_panes = 'A2'

# Identify all preorder orders and count preorder products first
preorder_oids = set()
for oid in all_order_ids:
    crm = crm_orders.get(oid, {})
    crm_approval = val(crm.get('Trạng thái phê duyệt', ''))
    crm_accounting = val(crm.get('Tình trạng ghi doanh số', ''))
    crm_payment = val(crm.get('Tình trạng thanh toán', ''))
    crm_delivery = val(crm.get('Tình trạng giao hàng', ''))
    
    is_crm_preorder = (
        crm_approval == 'Đã duyệt' and
        crm_accounting == 'Đề nghị ghi' and
        crm_payment in PAID and
        crm_delivery not in DELIVERY_DELIVERED
    )
    if is_crm_preorder or oid in pre_orders:
        preorder_oids.add(oid)

preorder_prod_count = 0
for oid in all_order_ids:
    crm = crm_orders.get(oid, {})
    misa = misa_orders.get(oid, {})
    
    # MISA sales
    misa_sales_by_prod = defaultdict(float)
    misa_invoices_by_prod = defaultdict(set)
    for r in so_orders.get(oid, []):
        mã_hàng = val(r.get('Mã hàng', ''))
        qty = float_val(r.get('Số lượng bán', 0))
        inv = val(r.get('Số hóa đơn', ''))
        misa_sales_by_prod[mã_hàng] += qty
        if inv:
            misa_invoices_by_prod[mã_hàng].add(inv)
            
    is_preorder = oid in preorder_oids
    processed_prods = set()
    
    for li in crm_lineitems.get(oid, []):
        ma_hang = val(li.get('Mã hàng hóa', ''))
        qty_crm = float_val(li.get('Số lượng', 0))
        qty_misa = misa_sales_by_prod.get(ma_hang, 0.0)
        invoices = misa_invoices_by_prod.get(ma_hang, set())
        qty_preorder = max(0.0, qty_crm - qty_misa)
        processed_prods.add(ma_hang)
        
        if is_preorder:
            if qty_preorder > 0 or (qty_misa > 0 and not invoices):
                preorder_prod_count += 1
        else:
            if qty_misa > 0 and not invoices:
                preorder_prod_count += 1
                
    for ma_hang, qty_misa in misa_sales_by_prod.items():
        if ma_hang in processed_prods:
            continue
        invoices = misa_invoices_by_prod.get(ma_hang, set())
        if qty_misa > 0 and not invoices:
            preorder_prod_count += 1

# ============================================================
# Sheet 0: Tổng quan (Dashboard)
# ============================================================
ws0 = wb_out.active
ws0.title = 'Tổng quan'

# Title
ws0.merge_cells('A1:H1')
cell = ws0['A1']
cell.value = 'BÁO CÁO ĐỐI SOÁT CRM ↔ MISA'
cell.font = Font(bold=True, size=16, color='2F5496')
cell.alignment = Alignment(horizontal='center')

ws0.merge_cells('A2:H2')
ws0['A2'].value = f'Ngày snapshot: {TODAY} | Tổng số đơn kiểm tra: {len(all_order_ids)}'
ws0['A2'].font = Font(size=11, color='666666')
ws0['A2'].alignment = Alignment(horizontal='center')

# Summary stats
r = 4
summary_data = [
    ('Tổng quan dữ liệu', '', ''),
    ('Nguồn CRM Danh sách', len(crm_orders), ''),
    ('Nguồn MISA Kế toán', len(misa_orders), ''),
    ('Nguồn Sổ chi tiết bán hàng', len(so_orders), ''),
    ('Nguồn Pre-order feedback', len(pre_orders), ''),
    ('  Đơn hàng Pre-order cần XL', len(preorder_oids), 'Chi tiết tại sheet Pre-order cần XL'),
    ('  Sản phẩm Pre-order cần XL', preorder_prod_count, 'Chi tiết tại sheet Sản phẩm Pre-order'),
    ('Master audit cũ', len(master_orders), ''),
    ('', '', ''),
    ('Tổng số lỗi phát hiện', len(discrepancies), ''),
    ('  HIGH', severity_counts.get('HIGH', 0), 'Cần xử lý ngay'),
    ('  MEDIUM', severity_counts.get('MEDIUM', 0), 'Cần kiểm tra'),
    ('  LOW', severity_counts.get('LOW', 0), 'Cần theo dõi'),
    ('', '', ''),
    ('Phân bố theo loại lỗi', '', ''),
]
for t, c in type_counts.most_common(20):
    summary_data.append((f'  {t}', c, ''))
    
for i, (label, s_val, note) in enumerate(summary_data):
    cell = ws0.cell(row=r+i, column=1, value=label)
    if not label.startswith(' '):
        cell.font = Font(bold=True, size=12)
    else:
        cell.font = Font(size=11)
    ws0.cell(row=r+i, column=2, value=s_val)
    ws0.cell(row=r+i, column=3, value=note)

ws0.column_dimensions['A'].width = 40
ws0.column_dimensions['B'].width = 15
ws0.column_dimensions['C'].width = 30

# ============================================================
# Shared Discrepancy Headers (used by all discrepancy sheets)
# ============================================================
headers = ['ID', 'Loại', 'Severity', 'Order ID', 'Khách hàng', 'Owner', 'Giai đoạn',
           'CRM Giao hàng', 'CRM Thanh toán', 'CRM Ghi DS', 'CRM Duyệt', 'CRM Thực hiện', 'CRM Xuất HĐ', 'CRM Giá trị HĐ',
           'MISA Giao hàng', 'MISA Thực thu', 'MISA Ghi DS', 'MISA Xuất HĐ', 'MISA Giá trị HĐ',
           'Chi tiết', 'Gợi ý xử lý', 'Đã xử lý', 'Nhóm cũ', 'Ghi chú cũ', 'Gợi ý sửa cũ', 'Nguồn file']

col_widths = [6, 24, 8, 16, 25, 20, 10, 14, 14, 18, 14, 14, 14, 14, 14, 14, 18, 14, 14, 50, 45, 15, 10, 40, 40, 20]

# Group discrepancies by order_id for quick lookup
discs_by_oid = defaultdict(list)
for d in discrepancies:
    discs_by_oid[d['order_id']].append(d)

# ============================================================
# Sheet 1: Tất cả đơn hàng (All reconciled orders)
# ============================================================
ws_all = wb_out.create_sheet('Tất cả đơn hàng')
all_orders_headers = [
    'Order ID', 'Khách hàng', 'Owner', 'Giai đoạn',
    'CRM Duyệt', 'CRM Giao hàng', 'CRM Thanh toán', 'CRM Ghi DS', 'CRM Thực hiện', 'CRM Xuất HĐ', 'CRM Giá trị HĐ',
    'MISA Giao hàng', 'MISA Thực thu', 'MISA Ghi DS', 'MISA Xuất HĐ', 'MISA Giá trị HĐ',
    'Lỗi đối soát', 'Chi tiết lỗi', 'Gợi ý xử lý', 'Mức độ lỗi',
    'Đã xử lý',
    'Nhóm cũ', 'Ghi chú cũ', 'Gợi ý sửa cũ', 'Nguồn dữ liệu'
]
all_orders_widths = [16, 25, 20, 10, 14, 14, 14, 18, 14, 14, 14, 14, 18, 18, 14, 14, 25, 50, 45, 12, 15, 10, 40, 40, 20]

all_orders_rows = []
for oid in sorted(all_order_ids):
    crm = crm_orders.get(oid, {})
    misa = misa_orders.get(oid, {})
    master = master_orders.get(oid, {})
    
    cust = crm.get('Khách hàng', '') or misa.get('Khách hàng', '') or master.get('Khách hàng', '') or ''
    owner = crm.get('Người thực hiện', '') or misa.get('Người thực hiện', '') or ''
    stage = crm.get('Giai đoạn', '')
    
    crm_app = crm.get('Trạng thái phê duyệt', '')
    crm_del = crm.get('Tình trạng giao hàng', '')
    crm_pay = crm.get('Tình trạng thanh toán', '')
    crm_acc = crm.get('Tình trạng ghi doanh số', '')
    crm_exec = crm.get('Tình trạng', '')
    crm_inv = crm.get('Đã xuất hóa đơn', '')
    crm_inv_v = float_val(crm.get('Giá trị đã xuất hóa đơn', ''))
    crm_inv_val_str = f'{crm_inv_v:,.0f}' if crm_inv_v else ''
    
    misa_del = misa.get('Tình trạng giao hàng', '')
    misa_thucthu = float_val(misa.get('Thực thu', ''))
    misa_pay_str = f'Thực thu={misa_thucthu:,.0f}' if misa_thucthu else ''
    misa_acc = misa.get('Tình trạng ghi doanh số', '')
    misa_inv = misa.get('Tình trạng xuất hóa đơn', '')
    misa_inv_v = float_val(misa.get('Giá trị đã xuất hóa đơn', ''))
    misa_inv_val_str = f'{misa_inv_v:,.0f}' if misa_inv_v else ''
    
    # Check if there are discrepancies
    oid_discs = discs_by_oid.get(oid, [])
    if oid_discs:
        err_types = ', '.join(d['type'] for d in oid_discs)
        err_detail = '; '.join(d['detail'] for d in oid_discs)
        sug_str = '; '.join(d['suggestion'] for d in oid_discs if d['suggestion'])
        severity = 'HIGH' if any(d['severity'] == 'HIGH' for d in oid_discs) else ('MEDIUM' if any(d['severity'] == 'MEDIUM' for d in oid_discs) else 'LOW')
    else:
        err_types = 'Không có lỗi'
        err_detail = 'Khớp hoàn toàn'
        sug_str = ''
        severity = 'OK'
        
    old_grp = master.get('primary_issue_group', '')
    old_cmt = master.get('review_comment', '')
    old_sug = master.get('suggested_fix', '')
    
    # Nguồn file
    sources = []
    if oid in crm_orders: sources.append('CRM_Sale')
    if oid in misa_orders: sources.append('MISA_Accounting')
    if oid in so_orders: sources.append('So_chi_tiet_ban_hang')
    if oid in pre_orders: sources.append('Pre-order')
    if oid in master_orders: sources.append('Master_audit')
    source_str = '+'.join(sources)
    
    all_orders_rows.append([
        oid, cust, owner, stage,
        crm_app, crm_del, crm_pay, crm_acc, crm_exec, crm_inv, crm_inv_val_str,
        misa_del, misa_pay_str, misa_acc, misa_inv, misa_inv_val_str,
        err_types, err_detail, sug_str, severity,
        check_processed_status(oid),
        old_grp, old_cmt, old_sug, source_str
    ])

write_sheet(ws_all, 'Tất cả đơn hàng', all_orders_headers, all_orders_rows, all_orders_widths)

# Color code ws_all based on severity
for r_idx, row in enumerate(all_orders_rows, 2):
    severity = row[19]
    fill = None
    if severity == 'HIGH':
        fill = red_fill
    elif severity == 'MEDIUM':
        fill = yellow_fill
    elif severity == 'LOW':
        fill = green_fill
    
    if fill:
        for c in range(1, len(all_orders_headers)+1):
            ws_all.cell(row=r_idx, column=c).fill = fill

# ============================================================
# Sheet 2: Quá hạn giao (Separated overdue delivery errors)
# ============================================================
ws_overdue = wb_out.create_sheet('Quá hạn giao')
overdue_types = {'G-Overdue', 'J-Paid_Not_Delivered', 'K-Old_G2_Unresolved'}
overdue_discs = [d for d in discrepancies if d['type'] in overdue_types]
overdue_rows = []
for d in overdue_discs:
    overdue_rows.append([
        d['id'], d['type'], d['severity'], d['order_id'], d['customer'], d['owner'], d['gioi_doan'],
        d['CRM_delivery'], d['CRM_payment'], d['CRM_accounting'], d['CRM_approval'], d['CRM_execution'], d['CRM_invoiced'], d['CRM_invoice_value'],
        d['MISA_delivery'], d['MISA_thuc_thu'], d['MISA_accounting'], d['MISA_invoiced'], d['MISA_invoice_value'],
        d['detail'], d['suggestion'], check_processed_status(d['order_id']), d['old_group'], d['old_comment'], d['old_suggested_fix'], d['source_files'],
    ])
write_sheet(ws_overdue, 'Quá hạn giao', headers, overdue_rows, col_widths)

for r_idx, d in enumerate(overdue_discs, 2):
    fill = yellow_fill
    if d['severity'] == 'HIGH':
        fill = red_fill
    elif d['severity'] == 'LOW':
        fill = green_fill
    for c in range(1, len(headers)+1):
        ws_overdue.cell(row=r_idx, column=c).fill = fill

# ============================================================
# Sheet 2: Pre-order cần XL
# ============================================================
ws_preorder = wb_out.create_sheet('Pre-order cần XL')

pre_headers = [
    'Order ID', 'Khách hàng', 'Owner', 
    'CRM Duyệt', 'CRM Giao hàng', 'CRM Thanh toán', 'CRM Ghi DS', 'CRM Thực hiện',
    'MISA Giao hàng', 'MISA Thực thu', 
    'Lỗi đối soát', 'Gợi ý xử lý', 'Đã xử lý', 'Note phản hồi',
    'Nhóm cũ', 'Ghi chú cũ', 'Gợi ý sửa cũ'
]

pre_rows = []
for oid in sorted(preorder_oids):
    order_discs = [d for d in discrepancies if d['order_id'] == oid]
    pre_info = pre_orders.get(oid, {})
    note_str = '; '.join(sorted(pre_info.get('note_actions', []))) if pre_info else ''
    
    if not order_discs and not note_str:
        continue
        
    crm = crm_orders.get(oid, {})
    misa = misa_orders.get(oid, {})
    
    # CRM values
    crm_cust = val(crm.get('Khách hàng', ''))
    crm_owner = val(crm.get('Người thực hiện', ''))
    crm_approval = val(crm.get('Trạng thái phê duyệt', ''))
    crm_del = val(crm.get('Tình trạng giao hàng', ''))
    crm_pay = val(crm.get('Tình trạng thanh toán', ''))
    crm_acc = val(crm.get('Tình trạng ghi doanh số', ''))
    crm_exec = val(crm.get('Tình trạng', ''))
    
    # MISA values
    misa_del = val(misa.get('Tình trạng giao hàng', ''))
    misa_thucthu = float_val(misa.get('Thực thu', ''))
    misa_thucthu_str = f'{misa_thucthu:,.0f}' if misa_thucthu else '0'
    
    # Combine errors and suggestions
    errs = []
    sugs = []
    for d in order_discs:
        errs.append(f"[{d['type']}] {d['detail']}")
        if d['suggestion'] and d['suggestion'] not in sugs:
            sugs.append(d['suggestion'])
            
    errors_str = '; '.join(errs)
    sug_str = '; '.join(sugs)
    
    master = master_orders.get(oid, {})
    old_grp = master.get('primary_issue_group', '')
    old_cmt = master.get('review_comment', '')
    old_sug = master.get('suggested_fix', '')
    
    pre_rows.append([
        oid, crm_cust, crm_owner,
        crm_approval, crm_del, crm_pay, crm_acc, crm_exec,
        misa_del, misa_thucthu_str,
        errors_str, sug_str, check_processed_status(oid), note_str,
        old_grp, old_cmt, old_sug
    ])

write_sheet(ws_preorder, 'Pre-order cần XL', pre_headers, pre_rows, [16, 25, 20, 14, 14, 14, 14, 14, 14, 14, 45, 45, 15, 30, 10, 40, 40])

# Color preorder rows
for r_idx, row in enumerate(pre_rows, 2):
    oid = row[0]
    order_discs = [d for d in discrepancies if d['order_id'] == oid]
    fill = None
    if any(d['severity'] == 'HIGH' for d in order_discs):
        fill = red_fill
    elif any(d['severity'] == 'MEDIUM' for d in order_discs):
        fill = yellow_fill
    elif any(d['severity'] == 'LOW' for d in order_discs):
        fill = green_fill
    if fill:
        for c in range(1, len(pre_headers)+1):
            ws_preorder.cell(row=r_idx, column=c).fill = fill

# ============================================================
# Sheet 3: Sản phẩm Pre-order (Product-level Pre-orders)
# ============================================================
ws_pre_prod = wb_out.create_sheet('Sản phẩm Pre-order')
prod_headers = [
    'Order ID', 'Trạng thái đơn hàng', 'Khách hàng', 'Owner', 'Mã hàng hóa', 'Tên hàng hóa',
    'SL CRM', 'Đơn giá', 'Thành tiền CRM', 
    'SL MISA bán', 'SL Pre-order còn lại', 'MISA Invoices', 'Trạng thái sản phẩm',
    'Danh sách hàng hóa trong đơn',
    'Đã xử lý',
    'Nhóm cũ', 'Ghi chú cũ', 'Gợi ý sửa cũ'
]
prod_widths = [16, 45, 25, 20, 15, 35, 10, 12, 14, 12, 18, 18, 45, 50, 15, 10, 40, 40]

prod_rows = []
for oid in sorted(all_order_ids):
    crm = crm_orders.get(oid, {})
    misa = misa_orders.get(oid, {})
    master = master_orders.get(oid, {})
    
    cust = crm.get('Khách hàng', '') or misa.get('Khách hàng', '') or master.get('Khách hàng', '') or ''
    owner = crm.get('Người thực hiện', '') or misa.get('Người thực hiện', '') or ''
    
    # Get statuses for description column
    crm_approval = val(crm.get('Trạng thái phê duyệt', ''))
    crm_accounting = val(crm.get('Tình trạng ghi doanh số', ''))
    crm_payment = val(crm.get('Tình trạng thanh toán', ''))
    crm_delivery = val(crm.get('Tình trạng giao hàng', ''))
    crm_execution = val(crm.get('Tình trạng', ''))
    
    misa_delivery = val(misa.get('Tình trạng giao hàng', ''))
    misa_payment_thucthu = float_val(misa.get('Thực thu', ''))
    misa_thucthu_str = f'{misa_payment_thucthu:,.0f}' if misa_payment_thucthu else '0'
    
    status_parts = []
    if oid in crm_orders:
        status_parts.append(f"CRM: {crm_approval}/{crm_accounting}/{crm_payment}/{crm_delivery or 'Trống'}/{crm_execution or 'Trống'}")
    else:
        status_parts.append("CRM: KHÔNG CÓ ĐƠN")
    if oid in misa_orders:
        status_parts.append(f"MISA: Giao={misa_delivery or 'Trống'}, Thu={misa_thucthu_str}")
    else:
        status_parts.append("MISA: KHÔNG CÓ ĐƠN")
    order_status_str = " | ".join(status_parts)
    
    # Build full list of goods in this order
    crm_goods_list = []
    for li in crm_lineitems.get(oid, []):
        g_code = val(li.get('Mã hàng hóa', ''))
        g_name = val(li.get('Diễn giải', '')) or val(li.get('Mô tả', ''))
        g_qty = float_val(li.get('Số lượng', 0))
        crm_goods_list.append(f"{g_code} ({g_name}) x {g_qty:,.0f}")
    
    if not crm_goods_list:
        misa_goods_list = []
        for r in so_orders.get(oid, []):
            g_code = val(r.get('Mã hàng', ''))
            g_name = val(r.get('Tên hàng', ''))
            g_qty = float_val(r.get('Số lượng bán', 0))
            misa_goods_list.append(f"{g_code} ({g_name}) x {g_qty:,.0f}")
        goods_in_order = '; '.join(misa_goods_list)
    else:
        goods_in_order = '; '.join(crm_goods_list)
    
    # MISA sales for this order by product code
    misa_sales_by_prod = defaultdict(float)
    misa_invoices_by_prod = defaultdict(set)
    misa_names_by_prod = {}
    misa_prices_by_prod = {}
    for r in so_orders.get(oid, []):
        mã_hàng = val(r.get('Mã hàng', ''))
        qty = float_val(r.get('Số lượng bán', 0))
        inv = val(r.get('Số hóa đơn', ''))
        misa_sales_by_prod[mã_hàng] += qty
        if inv:
            misa_invoices_by_prod[mã_hàng].add(inv)
        if mã_hàng not in misa_names_by_prod:
            misa_names_by_prod[mã_hàng] = val(r.get('Tên hàng', ''))
        if mã_hàng not in misa_prices_by_prod:
            misa_prices_by_prod[mã_hàng] = float_val(r.get('Đơn giá', 0))
            
    old_grp = master.get('primary_issue_group', '')
    old_cmt = master.get('review_comment', '')
    old_sug = master.get('suggested_fix', '')
    
    is_preorder = oid in preorder_oids
    processed_prods = set()
    
    # Process CRM line items
    for li in crm_lineitems.get(oid, []):
        ma_hang = val(li.get('Mã hàng hóa', ''))
        mo_ta = val(li.get('Mô tả', ''))
        qty_crm = float_val(li.get('Số lượng', 0))
        price = float_val(li.get('Đơn giá', 0))
        total_crm = float_val(li.get('Thành tiền', 0))
        
        qty_misa = misa_sales_by_prod.get(ma_hang, 0.0)
        invoices = misa_invoices_by_prod.get(ma_hang, set())
        
        qty_preorder = max(0.0, qty_crm - qty_misa)
        invoices_str = ', '.join(sorted(invoices))
        processed_prods.add(ma_hang)
        
        # Determine status and severity
        if is_preorder:
            if qty_preorder > 0:
                if qty_misa == 0:
                    prod_status = "Pre-order thực tế (Chưa ghi nhận bán hàng & Chưa xuất hóa đơn trên MISA)"
                    severity = "MEDIUM"
                else:
                    prod_status = f"Pre-order thực tế một phần (Đã bán {qty_misa:,.0f} SP, còn lại {qty_preorder:,.0f} SP chưa bán/HĐ)"
                    severity = "MEDIUM"
            elif qty_misa > 0 and not invoices:
                prod_status = "Đã ghi nhận bán hàng trên MISA nhưng chưa xuất hóa đơn (Cần kiểm tra/xuất HĐ)"
                severity = "HIGH"
            else:
                continue
        else:
            if qty_misa > 0 and not invoices:
                prod_status = "Đã ghi nhận bán hàng trên MISA nhưng chưa xuất hóa đơn (Cần kiểm tra/xuất HĐ)"
                severity = "HIGH"
            else:
                continue
            
        price_str = f'{price:,.0f}' if price else '0'
        total_crm_str = f'{total_crm:,.0f}' if total_crm else '0'
        
        prod_rows.append([
            oid, order_status_str, cust, owner, ma_hang, mo_ta,
            qty_crm, price_str, total_crm_str,
            qty_misa, qty_preorder, invoices_str, prod_status,
            goods_in_order,
            check_processed_status(oid),
            old_grp, old_cmt, old_sug, severity
        ])
        
    # Check any products sold in MISA but not in CRM
    for ma_hang, qty_misa in misa_sales_by_prod.items():
        if ma_hang in processed_prods:
            continue
        invoices = misa_invoices_by_prod.get(ma_hang, set())
        if qty_misa > 0 and not invoices:
            prod_status = "Đã ghi nhận bán hàng trên MISA nhưng chưa xuất hóa đơn (Cần kiểm tra/xuất HĐ)"
            severity = "HIGH"
            mo_ta = misa_names_by_prod.get(ma_hang, '')
            price = misa_prices_by_prod.get(ma_hang, 0.0)
            price_str = f'{price:,.0f}' if price else '0'
            total_crm_str = '0'
            invoices_str = ''
            
            prod_rows.append([
                oid, order_status_str, cust, owner, ma_hang, mo_ta,
                0.0, price_str, total_crm_str,
                qty_misa, 0.0, invoices_str, prod_status,
                goods_in_order,
                check_processed_status(oid),
                old_grp, old_cmt, old_sug, severity
            ])

write_sheet(ws_pre_prod, 'Sản phẩm Pre-order', prod_headers, [row[:-1] for row in prod_rows], prod_widths)

# Color preorder product rows
for r_idx, row in enumerate(prod_rows, 2):
    severity = row[-1]
    fill = None
    if severity == 'HIGH':
        fill = red_fill
    elif severity == 'MEDIUM':
        fill = yellow_fill
    elif severity == 'LOW':
        fill = green_fill
    if fill:
        for c in range(1, len(prod_headers)+1):
            ws_pre_prod.cell(row=r_idx, column=c).fill = fill

# ============================================================
# Sheet 4: Pre-order Full (All products in preorder orders)
# ============================================================
ws_pre_full = wb_out.create_sheet('Pre-order Full')
pre_full_rows = []

for oid in sorted(preorder_oids):
    crm = crm_orders.get(oid, {})
    misa = misa_orders.get(oid, {})
    master = master_orders.get(oid, {})
    
    cust = crm.get('Khách hàng', '') or misa.get('Khách hàng', '') or master.get('Khách hàng', '') or ''
    owner = crm.get('Người thực hiện', '') or misa.get('Người thực hiện', '') or ''
    
    # Get statuses for description column
    crm_approval = val(crm.get('Trạng thái phê duyệt', ''))
    crm_accounting = val(crm.get('Tình trạng ghi doanh số', ''))
    crm_payment = val(crm.get('Tình trạng thanh toán', ''))
    crm_delivery = val(crm.get('Tình trạng giao hàng', ''))
    crm_execution = val(crm.get('Tình trạng', ''))
    
    misa_delivery = val(misa.get('Tình trạng giao hàng', ''))
    misa_payment_thucthu = float_val(misa.get('Thực thu', ''))
    misa_thucthu_str = f'{misa_payment_thucthu:,.0f}' if misa_payment_thucthu else '0'
    
    status_parts = []
    if oid in crm_orders:
        status_parts.append(f"CRM: {crm_approval}/{crm_accounting}/{crm_payment}/{crm_delivery or 'Trống'}/{crm_execution or 'Trống'}")
    else:
        status_parts.append("CRM: KHÔNG CÓ ĐƠN")
    if oid in misa_orders:
        status_parts.append(f"MISA: Giao={misa_delivery or 'Trống'}, Thu={misa_thucthu_str}")
    else:
        status_parts.append("MISA: KHÔNG CÓ ĐƠN")
    order_status_str = " | ".join(status_parts)
    
    # Build full list of goods in this order
    crm_goods_list = []
    for li in crm_lineitems.get(oid, []):
        g_code = val(li.get('Mã hàng hóa', ''))
        g_name = val(li.get('Diễn giải', '')) or val(li.get('Mô tả', ''))
        g_qty = float_val(li.get('Số lượng', 0))
        crm_goods_list.append(f"{g_code} ({g_name}) x {g_qty:,.0f}")
    
    if not crm_goods_list:
        misa_goods_list = []
        for r in so_orders.get(oid, []):
            g_code = val(r.get('Mã hàng', ''))
            g_name = val(r.get('Tên hàng', ''))
            g_qty = float_val(r.get('Số lượng bán', 0))
            misa_goods_list.append(f"{g_code} ({g_name}) x {g_qty:,.0f}")
        goods_in_order = '; '.join(misa_goods_list)
    else:
        goods_in_order = '; '.join(crm_goods_list)
        
    misa_sales_by_prod = defaultdict(float)
    misa_invoices_by_prod = defaultdict(set)
    misa_names_by_prod = {}
    misa_prices_by_prod = {}
    for r in so_orders.get(oid, []):
        mã_hàng = val(r.get('Mã hàng', ''))
        qty = float_val(r.get('Số lượng bán', 0))
        inv = val(r.get('Số hóa đơn', ''))
        misa_sales_by_prod[mã_hàng] += qty
        if inv:
            misa_invoices_by_prod[mã_hàng].add(inv)
        if mã_hàng not in misa_names_by_prod:
            misa_names_by_prod[mã_hàng] = val(r.get('Tên hàng', ''))
        if mã_hàng not in misa_prices_by_prod:
            misa_prices_by_prod[mã_hàng] = float_val(r.get('Đơn giá', 0))
            
    old_grp = master.get('primary_issue_group', '')
    old_cmt = master.get('review_comment', '')
    old_sug = master.get('suggested_fix', '')
    
    processed_prods = set()
    
    # Process CRM line items
    for li in crm_lineitems.get(oid, []):
        ma_hang = val(li.get('Mã hàng hóa', ''))
        mo_ta = val(li.get('Mô tả', ''))
        qty_crm = float_val(li.get('Số lượng', 0))
        price = float_val(li.get('Đơn giá', 0))
        total_crm = float_val(li.get('Thành tiền', 0))
        
        qty_misa = misa_sales_by_prod.get(ma_hang, 0.0)
        invoices = misa_invoices_by_prod.get(ma_hang, set())
        
        qty_preorder = max(0.0, qty_crm - qty_misa)
        invoices_str = ', '.join(sorted(invoices))
        processed_prods.add(ma_hang)
        
        # Determine status and severity
        if qty_preorder > 0:
            if qty_misa == 0:
                prod_status = "Pre-order thực tế (Chưa ghi nhận bán hàng & Chưa xuất hóa đơn trên MISA)"
                severity = "MEDIUM"
            else:
                prod_status = f"Pre-order thực tế một phần (Đã bán {qty_misa:,.0f} SP, còn lại {qty_preorder:,.0f} SP chưa bán/HĐ)"
                severity = "MEDIUM"
        elif qty_misa > 0 and not invoices:
            prod_status = "Đã ghi nhận bán hàng trên MISA nhưng chưa xuất hóa đơn (Cần kiểm tra/xuất HĐ)"
            severity = "HIGH"
        else:
            prod_status = "Đã hoàn thành (Đã giao hàng & Xuất hóa đơn đủ)"
            severity = "OK"
            
        price_str = f'{price:,.0f}' if price else '0'
        total_crm_str = f'{total_crm:,.0f}' if total_crm else '0'
        
        pre_full_rows.append([
            oid, order_status_str, cust, owner, ma_hang, mo_ta,
            qty_crm, price_str, total_crm_str,
            qty_misa, qty_preorder, invoices_str, prod_status,
            goods_in_order,
            check_processed_status(oid),
            old_grp, old_cmt, old_sug, severity
        ])
        
    # Check any products sold in MISA but not in CRM
    for ma_hang, qty_misa in misa_sales_by_prod.items():
        if ma_hang in processed_prods:
            continue
        invoices = misa_invoices_by_prod.get(ma_hang, set())
        if qty_misa > 0:
            if not invoices:
                prod_status = "Đã ghi nhận bán hàng trên MISA nhưng chưa xuất hóa đơn (Cần kiểm tra/xuất HĐ)"
                severity = "HIGH"
            else:
                prod_status = "Đã bán bổ sung trên MISA (Đã xuất HĐ đầy đủ)"
                severity = "OK"
            mo_ta = misa_names_by_prod.get(ma_hang, '')
            price = misa_prices_by_prod.get(ma_hang, 0.0)
            price_str = f'{price:,.0f}' if price else '0'
            total_crm_str = '0'
            invoices_str = ', '.join(sorted(invoices))
            
            pre_full_rows.append([
                oid, order_status_str, cust, owner, ma_hang, mo_ta,
                0.0, price_str, total_crm_str,
                qty_misa, 0.0, invoices_str, prod_status,
                goods_in_order,
                check_processed_status(oid),
                old_grp, old_cmt, old_sug, severity
            ])

write_sheet(ws_pre_full, 'Pre-order Full', prod_headers, [row[:-1] for row in pre_full_rows], prod_widths)

# Color preorder full product rows
for r_idx, row in enumerate(pre_full_rows, 2):
    severity = row[-1]
    fill = None
    if severity == 'HIGH':
        fill = red_fill
    elif severity == 'MEDIUM':
        fill = yellow_fill
    elif severity == 'LOW':
        fill = green_fill
    if fill:
        for c in range(1, len(prod_headers)+1):
            ws_pre_full.cell(row=r_idx, column=c).fill = fill

# ============================================================
# Detailed Mismatch Sheets
# ============================================================
mismatch_groups = [
    {
        'title': 'Lệch hóa đơn (E2)',
        'types': {'E2-Invoice_Mismatch'}
    },
    {
        'title': 'NVGH chưa hoàn thành (I)',
        'types': {'I-NVGH_Activity_Open_Delivery'}
    },
    {
        'title': 'Chưa có hoạt động (N)',
        'types': {'N-No_Activity'}
    },
    {
        'title': 'Lệch thanh toán (C)',
        'types': {'C-Payment_Mismatch', 'C2-Payment_Mismatch'}
    },
    {
        'title': 'Thiếu đơn hàng (A)',
        'types': {'A-Missing_MISA', 'A2-Missing_CRM'}
    },
    {
        'title': 'Lỗi khác',
        'types': None
    }
]

captured_types = set()
for g in mismatch_groups:
    if g['types']:
        captured_types.update(g['types'])

for g in mismatch_groups:
    title = g['title']
    g_types = g['types']
    
    if g_types is not None:
        group_discs = [d for d in discrepancies if d['type'] in g_types and d['type'] not in overdue_types and d['order_id'] not in preorder_oids]
    else:
        group_discs = [d for d in discrepancies if d['type'] not in captured_types and d['type'] not in overdue_types and d['order_id'] not in preorder_oids]
        
    ws_g = wb_out.create_sheet(title)
    rows_data = []
    for d in group_discs:
        rows_data.append([
            d['id'], d['type'], d['severity'], d['order_id'], d['customer'], d['owner'], d['gioi_doan'],
            d['CRM_delivery'], d['CRM_payment'], d['CRM_accounting'], d['CRM_approval'], d['CRM_execution'], d['CRM_invoiced'], d['CRM_invoice_value'],
            d['MISA_delivery'], d['MISA_thuc_thu'], d['MISA_accounting'], d['MISA_invoiced'], d['MISA_invoice_value'],
            d['detail'], d['suggestion'], check_processed_status(d['order_id']), d['old_group'], d['old_comment'], d['old_suggested_fix'], d['source_files'],
        ])
        
    write_sheet(ws_g, title, headers, rows_data, col_widths)
    
    for r_idx, d in enumerate(group_discs, 2):
        fill = None
        if d['severity'] == 'HIGH':
            fill = red_fill
        elif d['severity'] == 'MEDIUM':
            fill = yellow_fill
        elif d['severity'] == 'LOW':
            fill = green_fill
        if fill:
            for c in range(1, len(headers)+1):
                ws_g.cell(row=r_idx, column=c).fill = fill

# ============================================================
# Save
# ============================================================
output_path = ROOT / f'bao_cao_doi_soat_CRM_MISA_{TODAY}.xlsx'
wb_out.save(output_path)
print(f"\n✅ Report saved to: {output_path}", file=sys.stderr)
print(f"   Total discrepancies: {len(discrepancies)}", file=sys.stderr)
print(f"   Sheets: {len(wb_out.sheetnames)}", file=sys.stderr)
