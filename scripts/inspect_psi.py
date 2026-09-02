import openpyxl

wb = openpyxl.load_workbook('test/PSI_Data 02.07.xlsx', data_only=True)
sheets = ['Pre-orders final', 'Revenue final', 'Sales CRM', 'Sales CRM (ĐH)']

for name in sheets:
    if name not in wb.sheetnames:
        print(f"Sheet {name} not found.")
        continue
    ws = wb[name]
    print(f"\n=== Sheet: {name} ===")
    
    # Print the first 15 rows to locate headers and sample data
    rows = list(ws.iter_rows(max_row=18, values_only=True))
    for idx, row in enumerate(rows, 1):
        # Only print non-completely empty rows
        if any(x is not None for x in row):
            # Print row index and the row values (truncated if long)
            print(f"Row {idx:02d}: {row[:12]}")
    print(f"Total rows: {ws.max_row}")
