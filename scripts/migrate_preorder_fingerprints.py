"""Migrate approved Pre-order exclusion fingerprints to stable Order + SKU keys."""

from pathlib import Path
from shutil import copy2
import sys

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

from psi_engine.manual_check import make_preorder_fingerprint


PATH = ROOT / "input/PSI_Manual_Check.xlsx"
BACKUP = ROOT / ".tmp/psi-20260807/PSI_Manual_Check.before_permanent_fingerprint.xlsx"


def main() -> None:
    BACKUP.parent.mkdir(parents=True, exist_ok=True)
    copy2(PATH, BACKUP)

    book = load_workbook(PATH)
    sheet = book["Preorder Exclusions"]
    headers = {cell.value: cell.column for cell in sheet[3] if cell.value}

    disposition_col = headers.get("Disposition")
    if disposition_col is None:
        disposition_col = sheet.max_column + 1
        sheet.cell(3, disposition_col).value = "Disposition"

    fingerprint_col = headers["Fingerprint"]
    status_col = headers["Status"]
    action_col = headers["Action"]
    order_col = headers["Order ID"]
    sku_col = headers["Canonical SKU"]

    changed = 0
    for row in range(4, sheet.max_row + 1):
        status = str(sheet.cell(row, status_col).value or "").strip().upper()
        action = str(sheet.cell(row, action_col).value or "").strip().upper()
        if status == "APPROVED" and action == "EXCLUDE FROM PREORDER":
            order_id = str(sheet.cell(row, order_col).value or "").strip()
            canonical_sku = str(sheet.cell(row, sku_col).value or "").strip()
            sheet.cell(row, fingerprint_col).value = make_preorder_fingerprint(
                action=action,
                order_id=order_id,
                canonical_sku=canonical_sku,
            )
            sheet.cell(row, disposition_col).value = "PERMANENT / DONE"
            changed += 1

    sheet.cell(2, 1).value = (
        f"{changed} dòng KT đã duyệt loại vĩnh viễn. "
        "Match theo ĐH + canonical SKU; Quantity + Net Value chỉ là snapshot audit."
    )
    book.save(PATH)
    print({"updated": changed, "path": str(PATH), "backup": str(BACKUP)})


if __name__ == "__main__":
    main()
