from pathlib import Path

from openpyxl import Workbook

FIXTURE_ROOT = Path(__file__).parent
CellValue = str | int | float | None


def save_book(name: str, rows: list[list[CellValue]], sheet: str = "Sheet") -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = sheet
    for row in rows:
        worksheet.append(row)
    workbook.save(FIXTURE_ROOT / name)


def padded_row(values: dict[int, CellValue], width: int = 80) -> list[CellValue]:
    row: list[CellValue] = [None] * width
    for index, value in values.items():
        row[index] = value
    return row


def main() -> None:
    save_book(
        "product_fixture.xlsx",
        [
            ["Mã hàng hóa", "Nguồn gốc", "Tên hàng", "Category", "Sub Category"],
            ["SKU-001", "Fixture Brand", "Known product", "Kitchen", "Tools"],
        ],
    )
    save_book(
        "malformed_product.xlsx",
        [["Mã hàng hóa"], ["SKU-001"]],
    )
    purchase_rows = [[None] * 80 for _ in range(3)]
    purchase_rows.append(
        padded_row(
            {
                1: "SỐ PO",
                2: "NGÀY PO",
                3: "SỐ DH",
                18: "MÃ MISA",
                19: "Tên hàng",
                20: "SL",
                26: "F.O.C",
                54: "NGÀY NHẬP KHO",
            }
        )
    )
    purchase_rows.extend(
        [
            padded_row({1: "PO-USED", 7: "NORMAL", 18: "SKU-001", 19: "Used item"}),
            padded_row({1: "PO-FOC", 7: "NORMAL", 18: "SKU-001", 19: "Free item", 26: "F.O.C"}),
            padded_row({1: "PO-MISSING", 7: "NORMAL", 19: "Missing MISA code"}),
        ]
    )
    save_book("loading_purchase_fixture.xlsx", purchase_rows, "LDL")
    revenue_rows = [[None] * 26 for _ in range(3)]
    revenue_rows.append(
        padded_row({10: "Mã hàng", 15: "Doanh số bán", 17: "Giá trị", 18: "TK Nợ", 19: "TK Có"}, 26)
    )
    revenue_rows.extend(
        [
            padded_row({10: "SKU-001", 11: "Known revenue", 12: "Revenue", 15: 3, 17: 300, 18: "5111"}, 26),
            padded_row({10: "SKU-MISSING", 11: "Missing revenue", 12: "Revenue", 15: 2, 17: 200, 19: "5112"}, 26),
            padded_row({10: "SKU-EXCLUDED", 11: "Excluded", 12: "KHÔNG PHẢI REVENUE", 15: 9, 17: 900, 18: "5111"}, 26),
        ]
    )
    save_book("so_chi_tiet_revenue_fixture.xlsx", revenue_rows)
    inventory_rows = [[None] * 15 for _ in range(5)]
    inventory_rows.append(padded_row({0: "Tên kho", 1: "Mã kho", 2: "Mã hàng", 11: "Cuối kỳ", 12: "Giá trị"}, 15))
    inventory_rows.extend(
        [
            padded_row({0: "KHO CHÍNH", 1: "MAIN", 2: "SKU-001", 11: 4, 12: 400}, 15),
            padded_row({0: "KHO BÌNH PHÚ (KHO LỖI)", 1: "BAD", 2: "SKU-001", 11: 8, 12: 800}, 15),
        ]
    )
    save_book("tong_hop_ton_kho_inventory_fixture.xlsx", inventory_rows)
    save_book("pre-order_fixture.xlsx", [["PRODUCT ID", "QUANTITY SOLD", "ĐH"], ["SKU-001", 1, "SO-1"]])
    crm = Workbook()
    orders = crm.active
    orders.title = "Danh sách"
    orders.append(["SALE ORDER", "Customer"])
    orders.append(["SO-1", "Fixture customer"])
    items = crm.create_sheet("Bảng hàng hóa")
    items.append(["PRODUCT ID", "SALE ORDER"])
    items.append(["SKU-001", "SO-1"])
    crm.save(FIXTURE_ROOT / "crm_sale_fixture.xlsx")
    save_book("target_fixture.xlsx", [["TARGET 2026"], [1000]])


if __name__ == "__main__":
    main()
