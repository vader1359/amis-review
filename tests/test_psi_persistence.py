import io
import json
import sys
import threading
import os
from datetime import date
from http.client import HTTPConnection
from pathlib import Path
from urllib.parse import urlencode

import pytest
from openpyxl import Workbook, load_workbook
from types import SimpleNamespace

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "web"))
import server  # noqa: E402
from test_psi_release_verifier import release_store
from psi_engine import PsiReleaseService, ReleaseConfig, ReleaseGateError, ReleaseRequest


FIXTURES = Path(__file__).parent / "fixtures"


def test_persistent_upload_keeps_snapshot_provenance() -> None:
    store = server.PsiMemoryStore()
    request = server.UploadRequest(
        team_id="team-a",
        actor_id="user-a",
        reporting_period="2026-07",
        data_as_of="2026-07-05",
        source_type="product",
        filename="product_fixture.xlsx",
        content= (FIXTURES / "product_fixture.xlsx").read_bytes(),
    )

    first = store.persist(request)
    second = store.persist(request)

    assert first.snapshot.object_path != second.snapshot.object_path
    assert first.snapshot.checksum_sha256 == second.snapshot.checksum_sha256
    assert first.run.fingerprint == second.run.fingerprint
    assert first.run.duplicate_policy == "active_mismatches_recorded"
    assert len(store.snapshots) == 2
    assert len(store.activity) == 2


def test_persistent_upload_rejects_wrong_actor_team() -> None:
    store = server.PsiMemoryStore()
    request = server.UploadRequest(
        team_id="team-a",
        actor_id="user-b",
        reporting_period="2026-07",
        data_as_of="2026-07-05",
        source_type="product",
        filename="product_fixture.xlsx",
        content=b"PK\x03\x04",
    )

    with pytest.raises(server.UploadAuthorizationError):
        store.persist(request)


def test_local_persist_accepts_upload_roles_without_auth_boundary() -> None:
    store = server.PsiMemoryStore(); server.H.store = store
    server.H.actors = {"viewer": ("user-a", "team-a"), "admin": ("admin", "team-a"), "contributor": ("user-a", "team-a")}
    server.H.roles = {"viewer": "viewer", "admin": "admin", "contributor": "contributor"}
    httpd = server.ThreadingHTTPServer(("127.0.0.1", 0), server.H); thread = threading.Thread(target=httpd.serve_forever, daemon=True); thread.start()
    try:
        content = (FIXTURES / "product_fixture.xlsx").read_bytes(); boundary = "----role-test"
        fields = {"team_id": "team-a", "reporting_period": "2026-07", "data_as_of": "2026-07-05", "source_type": "product"}
        parts = [f"--{boundary}\r\nContent-Disposition: form-data; name=\"{key}\"\r\n\r\n{value}\r\n".encode() for key, value in fields.items()]
        parts.append(f"--{boundary}\r\nContent-Disposition: form-data; name=\"file\"; filename=\"product_fixture.xlsx\"\r\nContent-Type: application/vnd.openxmlformats-officedocument.spreadsheetml.sheet\r\n\r\n".encode() + content + b"\r\n")
        body = b"".join(parts) + f"--{boundary}--\r\n".encode()
        for token, expected in (("viewer", 201), ("admin", 201), ("contributor", 201)):
            connection = HTTPConnection("127.0.0.1", httpd.server_address[1]); connection.request("POST", "/api/persist", body=body, headers={"Authorization": f"Bearer {token}", "Content-Type": f"multipart/form-data; boundary={boundary}"}); response = connection.getresponse(); response.read(); connection.close(); assert response.status == expected
    finally:
        httpd.shutdown(); httpd.server_close(); thread.join(timeout=2)


def test_authenticated_http_entrypoint_persists_upload_without_secret_response() -> None:
    store = server.PsiMemoryStore()
    server.H.store = store
    server.H.actors = {"test-token": ("user-a", "team-a")}
    server.H.roles = {"test-token": "contributor"}
    httpd = server.ThreadingHTTPServer(("127.0.0.1", 0), server.H)
    thread = threading.Thread(target=httpd.serve_forever, daemon=True)
    thread.start()
    try:
        content = (FIXTURES / "product_fixture.xlsx").read_bytes()
        boundary = "----psi-test"
        fields = {
            "team_id": "team-a",
            "reporting_period": "2026-07",
            "data_as_of": "2026-07-05",
            "source_type": "product",
        }
        parts = [
            f"--{boundary}\r\nContent-Disposition: form-data; name=\"{key}\"\r\n\r\n{value}\r\n".encode()
            for key, value in fields.items()
        ]
        parts.append(
            f"--{boundary}\r\nContent-Disposition: form-data; name=\"file\"; filename=\"product_fixture.xlsx\"\r\nContent-Type: application/vnd.openxmlformats-officedocument.spreadsheetml.sheet\r\n\r\n".encode()
            + content
            + b"\r\n"
        )
        body = b"".join(parts) + f"--{boundary}--\r\n".encode()
        connection = HTTPConnection("127.0.0.1", httpd.server_address[1])
        connection.request(
            "POST",
            "/api/persist",
            body=body,
            headers={
                "Authorization": "Bearer test-token",
                "Content-Type": f"multipart/form-data; boundary={boundary}",
            },
        )
        response = connection.getresponse()
        payload = json.loads(response.read())
        connection.close()
        assert response.status == 201
        assert payload["snapshot"]["checksum_sha256"]
        assert "service" not in json.dumps(payload).lower()
    finally:
        httpd.shutdown()
        httpd.server_close()
        thread.join(timeout=2)


def test_persistence_extracts_workbook_headers_and_row_count() -> None:
    store = server.PsiMemoryStore()
    request = server.UploadRequest(
        team_id="team-a", actor_id="user-a", reporting_period="2026-07",
        data_as_of="2026-07-05", source_type="product", filename="product_fixture.xlsx",
        content=(FIXTURES / "product_fixture.xlsx").read_bytes(),
    )

    persisted = store.persist(request)

    assert persisted.snapshot.header_preview[:4] == ("Mã hàng hóa", "Nguồn gốc", "Tên hàng", "Category")
    assert persisted.snapshot.row_count == 1


def test_persistence_maps_relational_entities_to_fake_repository() -> None:
    repository = server.PsiMemoryRepository()
    store = server.PsiMemoryStore(repository=repository)
    request = server.UploadRequest(
        team_id="team-a", actor_id="user-a", reporting_period="2026-07",
        data_as_of="2026-07-05", source_type="product", filename="product_fixture.xlsx",
        content=(FIXTURES / "product_fixture.xlsx").read_bytes(),
    )

    persisted = store.persist(request)

    assert repository.rows("upload_batches")
    assert repository.rows("source_snapshots")[0]["row_count"] == 1
    assert repository.rows("reconciliation_run_sources")[0]["source_snapshot_id"] == persisted.snapshot.id
    assert repository.rows("normalized_records")
    assert repository.rows("psi_drafts")[0]["reconciliation_run_id"] == persisted.run.id
    assert repository.rows("draft_sources")[0]["source_snapshot_id"] == persisted.snapshot.id
    assert repository.rows("activity_logs")


def test_new_mismatch_has_exact_location_and_handled_fingerprint_is_suppressed() -> None:
    repository = server.PsiMemoryRepository()
    store = server.PsiMemoryStore(repository=repository)
    store.persist(server.UploadRequest("team-a", "user-a", "2026-07", "2026-07-05", "product", "product_fixture.xlsx", (FIXTURES / "product_fixture.xlsx").read_bytes()))
    request = server.UploadRequest("team-a", "user-a", "2026-07", "2026-07-05", "revenue", "so_chi_tiet_revenue_fixture.xlsx", (FIXTURES / "so_chi_tiet_revenue_fixture.xlsx").read_bytes())
    store.persist(request)
    mismatches = repository.rows("mismatches")
    assert len(mismatches) == 1
    mismatch = mismatches[0]
    assert mismatch["status"] == "new"
    assert mismatch["values_by_source"] == {"file": "so_chi_tiet_revenue_fixture.xlsx", "sheet": "Sheet", "row": 6, "code": "SKU-MISSING", "description": "Missing revenue", "issue": "Missing Product mapping"}
    mismatch_id = str(mismatch["id"])
    repository.transition_mismatch(mismatch_id, "assigned", "", {}, "user-a")
    repository.transition_mismatch(mismatch_id, "in_progress", "", {}, "user-a")
    repository.transition_mismatch(mismatch_id, "known", "verified", {"source": "test"}, "user-a")
    workbook = load_workbook(io.BytesIO(request.content))
    workbook.active.insert_rows(5)
    shifted = io.BytesIO()
    workbook.save(shifted)
    workbook.close()
    store.persist(server.UploadRequest("team-a", "user-a", "2026-07", "2026-07-05", "revenue", "so_chi_tiet_revenue_fixture.xlsx", shifted.getvalue()))
    assert len(repository.rows("mismatches")) == 1
    assert repository.rows("mismatches")[0]["status"] == "known"


def test_invalid_calendar_and_malformed_multipart_are_controlled_4xx() -> None:
    store = server.PsiMemoryStore()
    request = server.UploadRequest(
        team_id="team-a", actor_id="user-a", reporting_period="2026-13",
        data_as_of="2026-02-31", source_type="product", filename="product_fixture.xlsx",
        content=(FIXTURES / "product_fixture.xlsx").read_bytes(),
    )
    with pytest.raises(server.UploadValidationError):
        store.persist(request)

    server.H.store = store
    server.H.actors = {"test-token": ("user-a", "team-a")}
    server.H.roles = {"test-token": "contributor"}
    httpd = server.ThreadingHTTPServer(("127.0.0.1", 0), server.H)
    thread = threading.Thread(target=httpd.serve_forever, daemon=True)
    thread.start()
    try:
        connection = HTTPConnection("127.0.0.1", httpd.server_address[1])
        connection.request("POST", "/api/persist", body=b"not multipart", headers={
            "Authorization": "Bearer test-token", "Content-Type": "application/json",
        })
        response = connection.getresponse()
        response.read()
        connection.close()
        assert response.status == 400
    finally:
        httpd.shutdown()
        httpd.server_close()
        thread.join(timeout=2)


def test_server_selects_memory_without_config_and_rejects_partial_supabase_config(monkeypatch: pytest.MonkeyPatch) -> None:
    monkeypatch.delenv("SUPABASE_URL", raising=False)
    monkeypatch.delenv("SUPABASE_SERVICE_ROLE_KEY", raising=False)
    assert isinstance(server.build_store(), server.PsiMemoryStore)
    monkeypatch.setenv("SUPABASE_URL", "http://127.0.0.1:9")
    with pytest.raises(server.UploadValidationError):
        server.build_store()


def test_authenticated_release_http_download_rejects_other_actor() -> None:
    store = release_store()
    server.H.store = store
    server.H.release_service = PsiReleaseService(store)
    server.H.actors = {"test-token": ("user-a", "team-a"), "other-token": ("user-b", "team-a")}
    server.H.roles = {"test-token": "admin", "other-token": "viewer"}
    httpd = server.ThreadingHTTPServer(("127.0.0.1", 0), server.H)
    thread = threading.Thread(target=httpd.serve_forever, daemon=True); thread.start()
    try:
        connection = HTTPConnection("127.0.0.1", httpd.server_address[1])
        connection.request("POST", "/api/release", body=b'{"reporting_period":"2026-07"}', headers={"Authorization": "Bearer test-token", "Content-Type": "application/json"})
        response = connection.getresponse(); payload = json.loads(response.read()); connection.close()
        assert response.status == 201
        token = payload["signed_url"].rsplit("/", 1)[1]
        connection = HTTPConnection("127.0.0.1", httpd.server_address[1])
        connection.request("GET", "/api/download/" + token, headers={"Authorization": "Bearer test-token"})
        response = connection.getresponse(); downloaded = response.read(); connection.close()
        assert response.status == 200
        load_workbook(io.BytesIO(downloaded), data_only=False).close()
        connection = HTTPConnection("127.0.0.1", httpd.server_address[1])
        connection.request("GET", "/api/download/" + token, headers={"Authorization": "Bearer other-token"})
        response = connection.getresponse(); response.read(); connection.close()
        assert response.status == 403
    finally:
        httpd.shutdown(); httpd.server_close(); thread.join(timeout=2)
