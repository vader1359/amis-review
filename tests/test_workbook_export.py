import sys
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
FIXTURES = Path(__file__).parent / "fixtures"
sys.path.insert(0, str(ROOT / "web"))
import server  # noqa: E402


EXPECTED_SHEETS = {
    "PSI Summary",
    "Mismatch",
    "Data gaps",
    "PO excluded",
    "PSI by Product",
    "Product detail",
    "Product final",
    "Purchase PO detail",
    "Purchase final",
    "Inventory source",
    "Inventory final",
    "Pre-order source",
    "Pre-orders final",
    "Revenue raw",
    "Revenue final",
    "CRM orders",
    "CRM items",
    "Target detail",
    "Target final",
}


def test_build_creates_openable_workbook_with_current_psi_sheets_when_fixture_set_is_processed(
    tmp_path: Path,
) -> None:
    # Given: the complete stable PSI fixture set.
    files = {
        path.name: path.read_bytes()
        for path in sorted(FIXTURES.glob("*.xlsx"))
        if path.name != "malformed_product.xlsx"
    }
    # When: the reconciliation build creates an XLSX artifact.
    result = server.build(files)
    workbook_path = tmp_path / "psi.xlsx"
    workbook_path.write_bytes(result["xlsx"])
    workbook = load_workbook(workbook_path, data_only=False)
    # Then: the artifact opens and exposes every current PSI sheet.
    assert EXPECTED_SHEETS <= set(workbook.sheetnames)
    assert workbook["PSI Summary"]["A1"].value == "Metric"
    assert workbook["Revenue final"]["AA2"].value == "=R2-MAX(0,U2)-X2-Y2"
